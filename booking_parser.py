"""
Booking Order Parser for MMG Backlite and Viola
Handles classification and parsing of booking order documents using OpenAI Responses API
"""

import json
import logging
import os
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime

import config
from pypdf import PdfMerger

logger = logging.getLogger("proposal-bot")

# File storage directories
BOOKING_ORDERS_BASE = Path("/data/booking_orders") if os.path.exists("/data") else Path(__file__).parent / "booking_orders"
# Combined PDFs directory (stores Excel + Original BO concatenated)
COMBINED_BOS_DIR = BOOKING_ORDERS_BASE / "combined_bos"
# Original uploaded BOs directory (permanent storage for user uploads during workflow)
ORIGINAL_BOS_DIR = BOOKING_ORDERS_BASE / "original_uploads"

# BO Template files (for future use - not currently used)
# These are the actual branded templates for Backlite and Viola
# Currently using simple Excel generation, will switch to template-based when ready
TEMPLATES_DIR = Path(__file__).parent / "bo_templates"
TEMPLATE_BACKLITE = TEMPLATES_DIR / "backlite_bo_template.xlsx"
TEMPLATE_VIOLA = TEMPLATES_DIR / "viola_bo_template.xlsx"

# Ensure directories exist
COMBINED_BOS_DIR.mkdir(parents=True, exist_ok=True)
ORIGINAL_BOS_DIR.mkdir(parents=True, exist_ok=True)


def sanitize_filename(filename: str) -> str:
    """
    Sanitize a string to be safe for use as a filename.
    Removes/replaces invalid characters for filesystems.

    Args:
        filename: The filename to sanitize

    Returns:
        Sanitized filename safe for filesystem use
    """
    import re
    # Remove or replace invalid characters: / \ : * ? " < > | and control characters
    sanitized = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', str(filename))
    # Replace multiple underscores/spaces with single underscore
    sanitized = re.sub(r'[_\s]+', '_', sanitized)
    # Remove leading/trailing underscores, dots, or spaces
    sanitized = sanitized.strip('_. ')
    # Limit length to 200 characters to be safe
    sanitized = sanitized[:200]
    # If empty after sanitization, use a default
    return sanitized if sanitized else 'booking_order'


@dataclass
class ParseResult:
    """Result from parsing a booking order"""
    data: Dict[str, Any]
    warnings: List[str]
    missing_required: List[str]
    needs_review: bool


class BookingOrderParser:
    """Parser for booking order documents with company-specific rules"""

    def __init__(self, company: str = "backlite"):
        self.company = company
        self.required_global = ["client", "net_pre_vat"]
        self.required_per_location = ["name", "start_date", "end_date", "net_amount"]

    def detect_file_type(self, file_path: Path) -> str:
        """Detect file type from extension"""
        suffix = file_path.suffix.lower()
        if suffix in [".xlsx", ".xls"]:
            return "excel"
        elif suffix == ".pdf":
            return "pdf"
        elif suffix in [".jpg", ".jpeg", ".png", ".gif", ".bmp"]:
            return "image"
        else:
            return "unknown"

    async def classify_document(self, file_path: Path, user_message: str = "", user_id: str = None) -> Dict[str, str]:
        """
        Classify document as BOOKING_ORDER or ARTWORK using OpenAI Responses API.
        Biased toward ARTWORK unless clear booking order fields present.

        Accepts any file type - will convert to PDF if needed for classification.

        Args:
            file_path: Path to the document file (any format)
            user_message: Optional user's message/prompt that accompanied the file upload
        """
        logger.info(f"[BOOKING PARSER] Classifying document: {file_path}")
        if user_message:
            logger.info(f"[BOOKING PARSER] User message context: {user_message}")

        # Convert to PDF if needed (OpenAI Responses API only accepts PDFs)
        pdf_path = file_path
        cleanup_pdf = False

        # Check file extension
        suffix = file_path.suffix.lower()

        # Convert images to PDF
        if suffix in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']:
            logger.info(f"[BOOKING PARSER] Converting image to PDF for classification...")
            try:
                from PIL import Image
                pdf_path = file_path.with_suffix('.pdf')
                img = Image.open(file_path)
                # Convert to RGB if needed (PNG with transparency, etc.)
                if img.mode in ('RGBA', 'LA', 'P'):
                    img = img.convert('RGB')
                img.save(pdf_path, 'PDF', resolution=100.0)
                cleanup_pdf = True
                logger.info(f"[BOOKING PARSER] Converted image to PDF: {pdf_path}")
            except Exception as e:
                logger.error(f"[BOOKING PARSER] Failed to convert image to PDF: {e}")
                return {"classification": "ARTWORK", "confidence": "high", "reasoning": "Image file detected"}

        # Convert Excel to PDF
        elif suffix in ['.xlsx', '.xls']:
            logger.info(f"[BOOKING PARSER] Converting Excel to PDF for classification...")
            try:
                pdf_path = await self._convert_excel_to_pdf(file_path)
                cleanup_pdf = True
                logger.info(f"[BOOKING PARSER] Converted Excel to PDF: {pdf_path}")
            except Exception as e:
                logger.error(f"[BOOKING PARSER] Failed to convert Excel to PDF: {e}")
                return {"classification": "UNKNOWN", "confidence": "low", "reasoning": f"Excel conversion failed: {e}"}

        # Upload PDF to OpenAI with purpose="user_data" (VendorAI pattern)
        try:
            with open(pdf_path, "rb") as f:
                file_obj = await config.openai_client.files.create(
                    file=f,
                    purpose="user_data"
                )
            file_id = file_obj.id
            logger.info(f"[BOOKING PARSER] Uploaded file to OpenAI: {file_id}")
        except Exception as e:
            logger.error(f"[BOOKING PARSER] Failed to upload file: {e}")
            if cleanup_pdf and pdf_path.exists():
                pdf_path.unlink(missing_ok=True)
            return {"classification": "UNKNOWN", "confidence": "low", "reasoning": str(e)}
        finally:
            # Clean up temporary PDF if we created one
            if cleanup_pdf and pdf_path.exists() and pdf_path != file_path:
                try:
                    pdf_path.unlink(missing_ok=True)
                except Exception as e:
                    logger.warning(f"[BOOKING PARSER] Failed to clean up temp PDF: {e}")

        # Classification prompt - BIASED toward ARTWORK
        user_context = f"\n\n**USER'S MESSAGE:** \"{user_message}\"\nUse this context to better understand the user's intent." if user_message else ""

        classification_prompt = f"""You are classifying a document as either a BOOKING_ORDER or ARTWORK.

**IMPORTANT BIAS:** Default to classifying as ARTWORK unless you see CLEAR booking order characteristics.

**BOOKING_ORDER characteristics (must have MULTIPLE of these):**
- Explicit booking/order reference numbers (e.g., "BO #12345", "Order #", "Booking Ref")
- Tables with columns: Location, Start Date, End Date, Duration, Costs
- Financial terms: Net, VAT, Gross, SLA deduction, Payment terms
- Client/Agency names with campaign details
- Multiple locations listed with dates and pricing

**ARTWORK characteristics (if you see ANY of these, classify as ARTWORK):**
- Visual designs, graphics, logos, imagery
- Product photos or promotional content
- Brand assets, mockups, creative designs
- Marketing materials
- No tabular pricing/booking data
{user_context}

**Classification rules:**
1. If you see ANY visual design elements → ARTWORK (high confidence)
2. If you see a table with locations/dates/pricing → BOOKING_ORDER (high confidence)
3. If unclear or minimal text → ARTWORK (medium confidence)
4. If it's a mix → ARTWORK (low confidence)
5. If user's message mentions "mockup", "billboard", "creative", "artwork" → ARTWORK (high confidence)
6. If user's message mentions "booking order", "BO", "parse" → BOOKING_ORDER (higher confidence)

**COMPANY DETECTION (for BOOKING_ORDER only):**
If classified as BOOKING_ORDER, determine the company:
- Look for "Backlite" or "BackLite" or "backlite" anywhere in document → company: "backlite"
- Look for "Viola" or "viola" anywhere in document → company: "viola"
- Check user's message for company name too
- If unclear, default to "backlite"

Analyze the uploaded file and respond with:
- classification: "BOOKING_ORDER" or "ARTWORK"
- confidence: "high", "medium", or "low"
- company: "backlite" or "viola" (ONLY if classification is BOOKING_ORDER, otherwise null)
- reasoning: Brief explanation (1 sentence)
"""

        try:
            # Use VendorAI syntax with structured JSON output
            response = await config.openai_client.responses.create(
                model=config.OPENAI_MODEL,
                input=[
                    {"role": "system", "content": "You are a document classifier. Analyze the file and provide classification in JSON format."},
                    {
                        "role": "user",
                        "content": [
                            {"type": "input_file", "file_id": file_id},
                            {"type": "input_text", "text": classification_prompt}
                        ]
                    }
                ],
                text={
                    "format": {
                        "type": "json_schema",
                        "name": "classification_response",
                        "strict": True,
                        "schema": {
                            "type": "object",
                            "properties": {
                                "classification": {
                                    "type": "string",
                                    "enum": ["BOOKING_ORDER", "ARTWORK"]
                                },
                                "confidence": {
                                    "type": "string",
                                    "enum": ["high", "medium", "low"]
                                },
                                "company": {
                                    "type": ["string", "null"],
                                    "enum": ["backlite", "viola", None]
                                },
                                "reasoning": {
                                    "type": "string"
                                }
                            },
                            "required": ["classification", "confidence", "company", "reasoning"],
                            "additionalProperties": False
                        }
                    }
                },
                store=False
            )

            # Track cost (user_id should already be converted to user_name by caller)
            import cost_tracking
            cost_tracking.track_openai_call(
                response=response,
                call_type="classification",
                workflow="bo_parsing",
                user_id=user_id,
                context=f"File: {file_path.name}",
                metadata={"file_type": suffix, "has_user_message": bool(user_message)}
            )

            if not response.output or len(response.output) == 0:
                logger.warning("[BOOKING PARSER] Empty classification response")
                return {"classification": "ARTWORK", "confidence": "low", "reasoning": "No response from model"}

            # Parse JSON response from structured output
            result_text = response.output_text
            logger.info(f"[BOOKING PARSER] Classification response: {result_text}")

            # Parse JSON directly (structured output guarantees valid JSON)
            result = json.loads(result_text)

            # Ensure company is set for BOOKING_ORDER (fallback to backlite if null)
            if result["classification"] == "BOOKING_ORDER" and not result.get("company"):
                result["company"] = "backlite"

            logger.info(f"[BOOKING PARSER] Parsed classification: {result}")
            return result

        except Exception as e:
            logger.error(f"[BOOKING PARSER] Classification error: {e}", exc_info=True)
            return {"classification": "ARTWORK", "confidence": "low", "reasoning": f"Error: {str(e)}"}
        finally:
            # Cleanup uploaded file
            try:
                await config.openai_client.files.delete(file_id)
            except:
                pass

    async def parse_file(self, file_path: Path, file_type: str, user_message: str = "", user_id: str = None) -> ParseResult:
        """
        Parse booking order file using OpenAI Responses API with structured JSON output.
        No hallucinations - only extract what's clearly present.

        Args:
            file_path: Path to the booking order file
            file_type: File type (pdf, image, excel)
            user_message: Optional user's message that accompanied the file upload
        """
        logger.info(f"[BOOKING PARSER] Parsing {file_type} file: {file_path}")
        if user_message:
            logger.info(f"[BOOKING PARSER] User message context: {user_message}")

        # Upload file to OpenAI with purpose="user_data" (VendorAI pattern)
        try:
            with open(file_path, "rb") as f:
                file_obj = await config.openai_client.files.create(
                    file=f,
                    purpose="user_data"
                )
            file_id = file_obj.id
            logger.info(f"[BOOKING PARSER] Uploaded file for parsing: {file_id}")
        except Exception as e:
            logger.error(f"[BOOKING PARSER] Failed to upload file for parsing: {e}")
            raise

        # Parsing prompt for structured extraction (no schema in prompt)
        parsing_prompt = self._build_parsing_prompt()

        # Add user message context if provided
        if user_message:
            parsing_prompt = f"""{parsing_prompt}

**USER'S MESSAGE CONTEXT:**
The user provided this message with the file: "{user_message}"

**CRITICAL - PRIORITIZE USER MESSAGE:** The user message often contains important information that may be missing from or clarify the document. Pay special attention to:
- **SLA amount/percentage**: User may specify "10% SLA" or "SLA is 5%" - ALWAYS extract this if mentioned
- **Payment terms**: User may specify "60 days PDC" or "30 days credit" - extract if mentioned
- **Sales person name**: User may say "this is for John" or "sales person is Sarah" - extract if mentioned
- **Location payment splits**: User may clarify "UAE02 is 100k, UAE03 is 200k" - use these exact amounts
- **Fee amounts**: User may specify "production fee is 2000" or "municipality fee is 500" - prioritize these values
- **Client/company details**: If user mentions client name or category, use that information
- **Any clarifications**: User message helps resolve ambiguities in the document

**PRIORITIZATION RULE:** If the user message contains explicit values for any field, those values OVERRIDE what's in the document.
"""

        try:
            # Use structured outputs with JSON schema + code_interpreter for better table parsing
            response = await config.openai_client.responses.create(
                model="gpt-5",
                reasoning={"effort": "high"},
                input=[
                    {"role": "system", "content": """You are a precise booking order data extractor.

**CRITICAL - MANDATORY FOR LOCATION TABLES AND FEES:**
1. **Carefully extract ALL table data** containing location names, dates, and amounts
2. **Pay special attention to tables with columns:** Location/Site, Start Date, End Date, Duration, Net Amount/Cost
3. **Also extract fee line items:** Production Fee, Upload Fee, Municipality Fee (DM), etc.
4. **Be precise with numbers** - only extract numbers that are clearly visible in the document
5. If any field is unclear or ambiguous, use null rather than guessing

**For other fields (client, brand, category):** Standard extraction is fine, reasonable inference allowed.
**For numbers, locations, and fees:** Extract exactly what you see. Do NOT estimate, interpolate, or calculate missing values.

**Fee Extraction Rules:**
- Look for rows labeled: "Production Fee", "Upload Fee", "Production Cost", "Upload Cost"
- Look for rows labeled: "DM Fee", "Municipality Fee", "Dubai Municipality"
- If fees are shown per-location in table, ADD them up into ONE global total
- If no fees found in document, use 0 (don't guess)"""},
                    {
                        "role": "user",
                        "content": [
                            {"type": "input_file", "file_id": file_id},
                            {"type": "input_text", "text": parsing_prompt}
                        ]
                    }
                ],
                # TODO: Re-enable code_interpreter after investigating timeout issues
                # tools=[
                #     {
                #         "type": "code_interpreter",
                #         "container": {
                #             "type": "auto"
                #         }
                #     }
                # ],
                text={
                    "format": {
                        "type": "json_schema",
                        "name": "booking_order_extraction",
                        "strict": True,
                        "schema": {
                            "type": "object",
                            "properties": {
                                "bo_number": {"type": ["string", "null"]},
                                "bo_date": {"type": ["string", "null"]},
                                "client": {"type": ["string", "null"]},
                                "agency": {"type": ["string", "null"]},
                                "brand_campaign": {"type": ["string", "null"]},
                                "category": {"type": ["string", "null"]},
                                "asset": {
                                    "anyOf": [
                                        {"type": "string"},
                                        {"type": "array", "items": {"type": "string"}},
                                        {"type": "null"}
                                    ]
                                },
                                "payment_terms": {"type": ["string", "null"]},
                                "sales_person": {"type": ["string", "null"]},
                                "commission_pct": {"type": ["number", "null"]},
                                "sla_pct": {"type": ["number", "null"]},
                                "municipality_fee": {"type": ["number", "null"]},
                                "production_upload_fee": {"type": ["number", "null"]},
                                "net_pre_vat": {"type": ["number", "null"]},
                                "vat_value": {"type": ["number", "null"]},
                                "gross_amount": {"type": ["number", "null"]},
                                "notes": {"type": ["string", "null"]},
                                "locations": {
                                    "type": "array",
                                    "items": {
                                        "type": "object",
                                        "properties": {
                                            "name": {"type": "string"},
                                            "asset": {"type": ["string", "null"]},
                                            "start_date": {"type": "string"},
                                            "end_date": {"type": "string"},
                                            "campaign_duration": {"type": "string"},
                                            "net_amount": {"type": "number"}
                                        },
                                        "required": ["name", "start_date", "end_date", "campaign_duration", "net_amount", "asset"],
                                        "additionalProperties": False
                                    }
                                }
                            },
                            "required": ["bo_number", "bo_date", "client", "agency", "brand_campaign", "category", "asset", "payment_terms", "sales_person", "commission_pct", "sla_pct", "municipality_fee", "production_upload_fee", "net_pre_vat", "vat_value", "gross_amount", "notes", "locations"],
                            "additionalProperties": False
                        }
                    }
                },
                store=False
            )

            # Track cost (user_id should already be converted to user_name by caller)
            import cost_tracking
            cost_tracking.track_openai_call(
                response=response,
                call_type="parsing",
                workflow="bo_parsing",
                user_id=user_id,
                context=f"File: {file_path.name}, Company: {self.company}",
                metadata={"file_type": file_type, "has_user_message": bool(user_message), "company": self.company}
            )

            if not response.output or len(response.output) == 0:
                raise ValueError("Empty parsing response from model")

            # Extract JSON from structured output using output_text
            # This gets the final text output after any tool calls
            result_text = response.output_text
            logger.info(f"[BOOKING PARSER] Parse response length: {len(result_text)} chars")
            logger.info(f"[BOOKING PARSER] Parse response text: {result_text[:500]}...")  # Log first 500 chars

            # Parse JSON (should be valid JSON from structured outputs)
            parsed_data = json.loads(result_text)

            # Post-process and validate
            processed = self._post_process_data(parsed_data)
            warnings = self._generate_warnings(processed)
            missing = self._check_missing_required(processed)
            needs_review = len(warnings) > 0 or len(missing) > 0

            return ParseResult(
                data=processed,
                warnings=warnings,
                missing_required=missing,
                needs_review=needs_review
            )

        except Exception as e:
            logger.error(f"[BOOKING PARSER] Parsing error: {e}", exc_info=True)
            raise
        finally:
            # Cleanup uploaded file
            try:
                await config.openai_client.files.delete(file_id)
            except:
                pass

    def _build_parsing_prompt(self) -> str:
        """Build the parsing prompt with field requirements - company-specific"""
        if self.company.lower() == "viola":
            return self._build_viola_parsing_prompt()
        else:
            return self._build_backlite_parsing_prompt()

    def _build_backlite_parsing_prompt(self) -> str:
        """Build Backlite-specific parsing prompt"""
        # Get static and digital locations for Backlite
        static_locations = []
        digital_locations = []
        for key, meta in config.LOCATION_METADATA.items():
            display_name = meta.get('display_name', key)
            if meta.get('display_type', '').lower() == 'static':
                static_locations.append(f"{display_name} ({key})")
            elif meta.get('display_type', '').lower() == 'digital':
                digital_locations.append(f"{display_name} ({key})")

        static_list = ", ".join(static_locations) if static_locations else "None"
        digital_list = ", ".join(digital_locations) if digital_locations else "None"

        location_context = f"""
**BACKLITE LOCATION REFERENCE (Use this to identify location types):**

NOTE: This list contains ALMOST ALL Backlite locations, but not necessarily every single one. If you see a location not in this list, use context clues from the BO to determine its type.

🔴 **DIGITAL LOCATIONS** (LED screens - get upload fees only):
{digital_list}

🔵 **STATIC LOCATIONS** (Traditional billboards - get production fees only):
{static_list}

Use this reference to determine if a location should have upload fees (digital) or production fees (static).
If a location isn't listed, make an intelligent guess based on naming patterns and fee descriptions in the BO.
"""

        return f"""You are an expert at extracting data from BACKLITE booking orders - a billboard/outdoor advertising company.

**TAKE YOUR TIME AND BE INTELLIGENT:**
These booking orders come from EXTERNAL clients and may have horrible, inconsistent structures. Do NOT rush. Carefully dissect the entire document, understand the business context, and intelligently parse the information. Think step-by-step about what you're seeing.
{location_context}
**CRITICAL BILLBOARD INDUSTRY CONTEXT:**

**Understanding Billboard Purchases:**
Booking orders (BOs) are contracts where clients purchase billboard advertising space. Key concepts:

1. **⚠️ CRITICAL: Bundled vs Separate Payments - DON'T DOUBLE COUNT!**

   **BUNDLED LOCATIONS (Split the payment):**
   When multiple locations appear TOGETHER in ONE row/line with ONE shared payment:
   - Example: "UAE02 & UAE03 - AED 320,000" → SPLIT 320k: UAE02=160k, UAE03=160k
   - Example: "Package (UAE03, UAE04, UAE05) - AED 480,000" → SPLIT: 160k each
   - Signals: Locations connected with "&", comma, or grouped in one table row
   - **DEFAULT BEHAVIOR:** If you see ONE total payment for ALL locations with no individual amounts specified, SPLIT IT EVENLY across all locations
   - Example: Document shows 5 locations with total payment of AED 500,000 and no per-location breakdown → Split evenly: 100k each

   **SEPARATE LOCATIONS (Full payment):**
   When a location has its own dedicated row with its own payment:
   - Example: "UAE21 - AED 120,000" on separate line → UAE21 gets full 120k

   **MIXED SCENARIO (Most common!):**
   ```
   Row 1: UAE02 & UAE03 - AED 320,000  → Split: UAE02=160k, UAE03=160k
   Row 2: UAE21 - AED 120,000          → Separate: UAE21=120k
   Total check: 160k + 160k + 120k = 440k ✅
   ```

   **❌ WRONG (Double counting):**
   ```
   UAE02=320k, UAE03=320k, UAE21=120k = 760k
   This is WRONG because UAE02 & UAE03 share the 320k!
   ```

   **How to identify bundled payments:**
   - Multiple location codes in same row (UAE02 & UAE03)
   - Package name covering multiple locations
   - Locations with "&" or "/" between them
   - Table structure: one payment cell spanning multiple location cells

2. **Fee Types (NOT locations - extract separately):**
   - **Municipality Fee:** Applies to ALL locations as a regulatory fee. Extract this as a separate global fee, NOT as a location
   - **Upload Fee:** Only for DIGITAL billboards (screens). This is the cost to upload creative content. NOT a location.
   - **Production Fee:** Only for STATIC billboards (printed). This is the cost to produce/print the creative. NOT a location.
   - These fees may appear mixed in with location rows - use intelligence to identify them

3. **SLA (Service Level Agreement) %:**
   - This is a DEDUCTION percentage (e.g., "0.4%" = 0.004 as decimal, "10%" = 0.10 as decimal)
   - Usually appears as a percentage field or row in the BO (look for "SLA", "SLA%", "Service Level Agreement")
   - **CRITICAL:** If you see an "SLA" field or row with a percentage, ALWAYS extract it! Don't leave it as null/0
   - **CHECK USER MESSAGE FIRST:** User often provides SLA in their message (e.g., "10% SLA", "SLA is 5%") - this OVERRIDES the document
   - Applied ONLY to the net rental amount (rental + production/upload fees + municipality fees)
   - SLA deduction happens BEFORE VAT is calculated
   - Formula: Net after SLA = Net rental - (Net rental × SLA%)
   - Then VAT is applied: Gross = (Net after SLA) × 1.05

4. **Location Types:**
   - **Digital/LED screens:** Get upload fees, no production fees
   - **Static billboards:** Get production fees, no upload fees
   - The document may not explicitly label these - use context clues

**FINANCIAL CALCULATION FLOW (CRITICAL - USE THIS TO IDENTIFY FEES):**

The standard calculation flow in billboard BOs is:
1. **Net Rental Amount** = Sum of all location rentals (e.g., AED 460,000)
2. **+ Production/Upload Fee** = Fee for creative production/upload (e.g., AED 2,000)
3. **+ Municipality Fee (DM Fee)** = Dubai Municipality regulatory fee (e.g., AED 520)
4. **= Net Amount** = Rental + Production + DM (e.g., AED 462,520)
5. **- SLA Deduction** = Net Amount × SLA% (e.g., 462,520 × 10% = AED 46,252)
6. **= Net Rental (after SLA)** = Net Amount - SLA (e.g., AED 414,000)
7. **+ VAT (5%)** = Usually applied to (Rental + Production), NOT DM (e.g., 5% of 462,000 = AED 23,126)
8. **= Gross Total** = Net Rental + VAT (e.g., AED 485,646)

**CRITICAL: How to identify Production/Upload Fees:**
- Look for line items labeled: "Production Fee", "Upload Fee", "Production Cost", "Upload Cost", "Creative Fee", "Net Production fee"
- Production/Upload fees are typically MEDIUM amounts (thousands, e.g., AED 1,000-5,000 per location or total)
- May appear in the costs table as a separate row, or in a summary section
- For DIGITAL screens: called "Upload Fee"
- For STATIC billboards: called "Production Fee" or "Production Cost"
- **IMPORTANT:** If shown per-location, ADD them ALL up into ONE global total
- If document shows "Production: AED 2,000" or "Upload: AED 1,500", extract this value
- If unclear or not found, use 0 (don't guess)

**CRITICAL: How to identify Municipality Fee:**
- Look for line items labeled: "DM Fee", "Municipality Fee", "Dubai Municipality", "Govt Fee"
- Municipality fee is typically a SMALL amount (hundreds to low thousands, not tens of thousands)
- If you see a breakdown showing: Rental + Production + Small Fee = Net Amount, that small fee is likely Municipality
- The Net Amount shown often equals: Rental + Production/Upload + Municipality
- Municipality fee may appear in a separate row in the costs table, or in a summary section

**Example:**
If you see:
- Net Rental amount: AED 460,000
- Net Production fee: AED 2,000
- Net DM fee: AED 520
- Net amount: AED 462,520 ← This confirms the municipality fee!
- VAT: AED 23,126
- Gross: AED 485,646

This tells you the municipality fee is AED 520.

**EXTRACTION RULES:**

**Be Intelligent About:**
- **Bundled payments:** If one payment covers multiple locations, split it across them
- **Fee identification:** Distinguish between location rentals vs fees (municipality/upload/production)
- **Inconsistent structures:** Tables may be messy, merged cells, unclear headers - parse carefully
- **Missing data:** Use null for truly missing fields, but try hard to find the data first
- **Calculations:** Extract what's shown, but understand the calculation logic to catch errors

**Format Rules:**
1. Dates: Convert to YYYY-MM-DD format (e.g., "21st Feb 2025" → "2025-02-21")
2. Percentages: Convert to decimal (e.g., "5%" → 0.05, "0.4%" → 0.004)
3. Numbers: Pure numbers without currency symbols (e.g., "AED 295,596.00" → 295596.00)
4. Asset codes: Extract as list: ["UAE02", "UAE03", "UAE21"]

**DOCUMENT STRUCTURE TO LOOK FOR:**

**Header Information:**
- BO Number (Booking Order reference - usually "BO-XXX" or "DPD-XXX")
- BO Date (when the BO was created)
- Client (company name purchasing the advertising)
  **IMPORTANT:** Client is the BUYER, NOT the seller:
  - Client = Company buying billboard advertising (e.g., "Emaar Properties", "Nestlé", "Mercedes-Benz")
  - DO NOT extract "Backlite" or "Viola" as the client - these are the SERVICE PROVIDERS (sellers)
  - Look for "From:", "Client:", "Advertiser:", or the company requesting the campaign
  - Client is who is PAYING for the billboard space, not who is selling it
- Agency (advertising agency, may be blank)
- Brand/Campaign (the advertised brand or campaign name)
  **IMPORTANT:** Use intelligent inference if brand/campaign is not explicitly stated:
  - If client is "Gucci LLC" → brand is likely "Gucci"
  - If client is "Emaar Properties PJSC" → brand is likely "Emaar"
  - If client is "Dubai Properties Development L.L.C" → brand is likely "Dubai Properties"
  - Extract the core brand name from the client company name by removing corporate suffixes like LLC, PJSC, L.L.C, Inc, Ltd, etc.
  - Only use the full client name as brand if there's truly no brand information anywhere in the document
- Category (the client's main industry/sector)
  **IMPORTANT:** Category represents the CLIENT'S industry, not the campaign type:
  - If client is "Emaar Properties PJSC" → category is "Real Estate"
  - If client is "Nestlé Middle East" → category is "FMCG" (Fast-Moving Consumer Goods)
  - If client is "Mercedes-Benz UAE" → category is "Automotive"
  - If client is "Emirates NBD" → category is "Banking/Finance"
  - Common categories: Real Estate, FMCG, Automotive, Banking/Finance, Hospitality, Retail, Healthcare, Technology, Entertainment
  - Infer from the client company name if not explicitly stated in the BO

**Location/Asset Details (usually in a table):**
For EACH billboard location, extract:
- Location name/code: **EXTRACT AS NATURAL LANGUAGE DISPLAY NAME**
  - ✅ CORRECT: "The Dubai Gateway", "Dubai Jawhara", "The Dubai Frame", "UAE02", "UAE03"
  - ❌ WRONG: "dubai_gateway", "dubai_jawhara", "dubai_frame" (these are system keys, not names!)
  - ❌ WRONG: "UAE02 (Unipole 16x8, Jebel Ali)", "The Gateway (LED)" (remove technical descriptions)
  - **Use the natural language name from the location reference list above** (e.g., "The Dubai Gateway" not "dubai_gateway")
  - Remove any parenthetical descriptions, dimensions, area names, or technical specs
  - If document shows "The Dubai Gateway (LED Screen)" → extract as "The Dubai Gateway"
  - If document shows "UAE02 (Unipole 16x8, Jebel Ali) & UAE03 (Billboard, Al Quoz)" → extract as TWO locations: "UAE02" and "UAE03"
  - Match the location to the display name from the reference list above when possible
- Start date (campaign start date)
- End date (campaign end date)
- Campaign duration **IMPORTANT:** Calculate and format intelligently:
  - Calculate number of days between start and end date
  - **Format rules (approximate to nearest unit):**
    * 28-31 days → "1 month"
    * 56-62 days → "2 months"
    * 84-93 days → "3 months"
    * 14-15 days → "2 weeks"
    * 21-22 days → "3 weeks"
    * 7-8 days → "1 week"
    * Only use "X days" if it doesn't fit these approximations
  - Examples: 30 days = "1 month", 15 days = "2 weeks", 60 days = "2 months", 10 days = "10 days"
  - If explicitly stated in BO (rare), use that value; otherwise calculate from dates
- Net amount (rental cost for THIS location - may need to split bundled payments)
- Production/Upload cost (if specified per-location)
- Type (digital vs static - if mentioned)

**CRITICAL: Understanding the Data Structure**

Booking orders have TWO types of costs:

1. **Per-Location Rental Amounts** (in locations array):
   - Each location has its own rental amount (e.g., UAE02: AED 80,000, UAE03: AED 80,000)
   - This is ONLY the rental cost for that specific billboard
   - Extract as `net_amount` for each location
   - **IMPORTANT:** If document shows ONLY a total rental amount with no per-location breakdown, SPLIT IT EVENLY across all locations
   - Example: 5 locations with total rental AED 500,000 → Each location gets AED 100,000

2. **GLOBAL Fees** (top-level fields, NOT per-location):
   - **municipality_fee**: Dubai Municipality regulatory fee for the ENTIRE booking
     - Look for: "DM Fee", "Municipality Fee", "Dubai Municipality", "Net DM fee"
     - Typically a small amount (hundreds to low thousands)
     - ONE total for all locations combined
     - **If document shows per-location DM fees:** ADD them all up into ONE total
     - Example: UAE02: AED 200, UAE03: AED 200, UAE21: AED 120 → municipality_fee: 520

   - **production_upload_fee**: Total production/upload cost for ALL locations
     - Look for: "Production Fee", "Upload Fee", "Production Cost", "Net Production fee"
     - This is production fee (for static) + upload fee (for digital) combined into ONE total
     - ONE total for all locations combined
     - **If document shows per-location fees:** ADD them all up into ONE total
     - Example: UAE02 upload: AED 500, UAE03 upload: AED 500, UAE21 production: AED 1000 → production_upload_fee: 2000

**CRITICAL:** Each location only stores its rental amount. ALL fees are GLOBAL single totals, NOT per-location values.
Even if the source document lists fees per location, you MUST sum them into single global totals.

**The Math:**
- Sum of all location rental amounts (e.g., 80,000 + 80,000 + 120,000 = 280,000)
- + production_upload_fee (e.g., 2,000)
- + municipality_fee (e.g., 520)
- = net_pre_vat (e.g., 282,520)

**Financial Totals:**
- Net amount (Rental + Production + Municipality, before SLA and VAT)
- VAT amount (5% tax, usually on Rental + Production, not Municipality)
- Gross amount (total including VAT)
- SLA % (deduction percentage, e.g., 10% = 0.10)

**Additional Information:**
- Payment terms (e.g., "60 days PDC", "30 days credit") - **CHECK USER MESSAGE** as this is often provided there
- Salesperson name - **CHECK USER MESSAGE** as user may specify who the sales person is
- Commission %

**CRITICAL NOTES:**
- If you see "Municipality Fee" or "Upload Fee" in the locations table, extract it separately as a fee, NOT as a location
- When multiple locations share one payment, split the amount across them intelligently
- Think about whether each line item is a location rental or a fee
- The structure may be messy - take your time to understand it
- Extract what you see, but understand the business logic to validate your extraction"""

    def _build_viola_parsing_prompt(self) -> str:
        """Build Viola-specific parsing prompt"""
        # TODO: Add Viola locations here when available
        viola_locations = "01A, 01B, 02A, 02B, 02C, 03A, 03B, 04A, 04B, 05A, 05B, 06A, 06B, 07A, 07B, 08A, 08B, 09A, 09B, 10A, 10B, 11A, 11B, 12A, 12B, 13A, 13B, 14A, 14B, 15A, 15B, 15C"

        return f"""You are an expert at extracting data from VIOLA booking orders - a billboard/outdoor advertising company.

**TAKE YOUR TIME AND BE INTELLIGENT:**
These booking orders come from EXTERNAL clients and may have horrible, inconsistent structures. Do NOT rush. Carefully dissect the entire document, understand the business context, and intelligently parse the information. Think step-by-step about what you're seeing.

**VIOLA LOCATION CODES:**
Viola uses alphanumeric location codes. Common codes include: {viola_locations}

**CRITICAL FOR VIOLA LOCATIONS:**
- Extract ONLY the location CODE (e.g., "04B", "03A", "15C")
- Viola locations often have descriptive names, but we only want the CODE
- Examples:
  - If document shows "Viola 04B - Sheikh Zayed Road Tower" → extract "04B"
  - If document shows "03A Al Barsha Mall" → extract "03A"
  - If document shows "Location 15C - Dubai Marina" → extract "15C"
- Look for the alphanumeric pattern (2 digits + optional letter)
- Strip away any descriptive text, addresses, or area names

**CRITICAL BILLBOARD INDUSTRY CONTEXT:**

**Understanding Billboard Purchases:**
Booking orders (BOs) are contracts where clients purchase billboard advertising space. Key concepts:

1. **⚠️ CRITICAL: Bundled vs Separate Payments - DON'T DOUBLE COUNT!**

   **BUNDLED LOCATIONS (Split the payment):**
   When multiple locations appear TOGETHER in ONE row/line with ONE shared payment:
   - Example: "04B & 03A - AED 320,000" → SPLIT 320k: 04B=160k, 03A=160k
   - Example: "Package (03A, 03B, 04A) - AED 480,000" → SPLIT: 160k each
   - Signals: Locations connected with "&", comma, or grouped in one table row

   **SEPARATE LOCATIONS (Full payment):**
   When a location has its own dedicated row with its own payment:
   - Example: "15C - AED 120,000" on separate line → 15C gets full 120k

   **MIXED SCENARIO (Most common!):**
   ```
   Row 1: 04B & 03A - AED 320,000  → Split: 04B=160k, 03A=160k
   Row 2: 15C - AED 120,000        → Separate: 15C=120k
   Total check: 160k + 160k + 120k = 440k ✅
   ```

   **❌ WRONG (Double counting):**
   ```
   04B=320k, 03A=320k, 15C=120k = 760k
   This is WRONG because 04B & 03A share the 320k!
   ```

   **How to identify bundled payments:**
   - Multiple location codes in same row (04B & 03A)
   - Package name covering multiple locations
   - Locations with "&" or "/" between them
   - Table structure: one payment cell spanning multiple location cells

2. **Fee Types (NOT locations - extract separately):**
   - **Municipality Fee:** Applies to ALL locations as a regulatory fee. Extract this as a separate global fee, NOT as a location
   - **Upload Fee:** For digital screens. This is the cost to upload creative content. NOT a location.
   - **Production Fee:** For static billboards. This is the cost to produce/print the creative. NOT a location.
   - These fees may appear mixed in with location rows - use intelligence to identify them

3. **SLA (Service Level Agreement) %:**
   - This is a DEDUCTION percentage (e.g., "0.4%" = 0.004 as decimal, "10%" = 0.10 as decimal)
   - Usually appears as a percentage field or row in the BO (look for "SLA", "SLA%", "Service Level Agreement")
   - **CRITICAL:** If you see an "SLA" field or row with a percentage, ALWAYS extract it! Don't leave it as null/0
   - **CHECK USER MESSAGE FIRST:** User often provides SLA in their message (e.g., "10% SLA", "SLA is 5%") - this OVERRIDES the document
   - Applied ONLY to the net rental amount (rental + production/upload fees + municipality fees)
   - SLA deduction happens BEFORE VAT is calculated
   - Formula: Net after SLA = Net rental - (Net rental × SLA%)
   - Then VAT is applied: Gross = (Net after SLA) × 1.05

**FINANCIAL CALCULATION FLOW (CRITICAL - USE THIS TO IDENTIFY FEES):**

The standard calculation flow in billboard BOs is:
1. **Net Rental Amount** = Sum of all location rentals (e.g., AED 460,000)
2. **+ Production/Upload Fee** = Fee for creative production/upload (e.g., AED 2,000)
3. **+ Municipality Fee (DM Fee)** = Dubai Municipality regulatory fee (e.g., AED 520)
4. **= Net Amount** = Rental + Production + DM (e.g., AED 462,520)
5. **- SLA Deduction** = Net Amount × SLA% (e.g., 462,520 × 10% = AED 46,252)
6. **= Net Rental (after SLA)** = Net Amount - SLA (e.g., AED 414,000)
7. **+ VAT (5%)** = Usually applied to (Rental + Production), NOT DM (e.g., 5% of 462,000 = AED 23,126)
8. **= Gross Total** = Net Rental + VAT (e.g., AED 485,646)

**CRITICAL: How to identify Production/Upload Fees:**
- Look for line items labeled: "Production Fee", "Upload Fee", "Production Cost", "Upload Cost", "Creative Fee", "Net Production fee"
- Production/Upload fees are typically MEDIUM amounts (thousands, e.g., AED 1,000-5,000 per location or total)
- May appear in the costs table as a separate row, or in a summary section
- **IMPORTANT:** If shown per-location, ADD them ALL up into ONE global total
- If document shows "Production: AED 2,000" or "Upload: AED 1,500", extract this value
- If unclear or not found, use 0 (don't guess)

**CRITICAL: How to identify Municipality Fee:**
- Look for line items labeled: "DM Fee", "Municipality Fee", "Dubai Municipality", "Govt Fee"
- Municipality fee is typically a SMALL amount (hundreds to low thousands, not tens of thousands)
- If you see a breakdown showing: Rental + Production + Small Fee = Net Amount, that small fee is likely Municipality
- The Net Amount shown often equals: Rental + Production/Upload + Municipality
- Municipality fee may appear in a separate row in the costs table, or in a summary section

**Example:**
If you see:
- Net Rental amount: AED 460,000
- Net Production fee: AED 2,000
- Net DM fee: AED 520
- Net amount: AED 462,520 ← This confirms the municipality fee!
- VAT: AED 23,126
- Gross: AED 485,646

This tells you the municipality fee is AED 520.

**EXTRACTION RULES:**

**Be Intelligent About:**
- **Bundled payments:** If one payment covers multiple locations, split it across them
- **Fee identification:** Distinguish between location rentals vs fees (municipality/upload/production)
- **Inconsistent structures:** Tables may be messy, merged cells, unclear headers - parse carefully
- **Missing data:** Use null for truly missing fields, but try hard to find the data first
- **Calculations:** Extract what's shown, but understand the calculation logic to catch errors

**Format Rules:**
1. Dates: Convert to YYYY-MM-DD format (e.g., "21st Feb 2025" → "2025-02-21")
2. Percentages: Convert to decimal (e.g., "5%" → 0.05, "0.4%" → 0.004)
3. Numbers: Pure numbers without currency symbols (e.g., "AED 295,596.00" → 295596.00)
4. Location codes: Extract as list: ["04B", "03A", "15C"] - CODES ONLY, no descriptive names

**DOCUMENT STRUCTURE TO LOOK FOR:**

**Header Information:**
- BO Number (Booking Order reference - usually "BO-XXX" or similar)
- BO Date (when the BO was created)
- Client (company name purchasing the advertising)
  **IMPORTANT:** Client is the BUYER, NOT the seller:
  - Client = Company buying billboard advertising (e.g., "Emaar Properties", "Nestlé", "Mercedes-Benz")
  - DO NOT extract "Viola" as the client - Viola is the SERVICE PROVIDER (seller)
  - Look for "From:", "Client:", "Advertiser:", or the company requesting the campaign
  - Client is who is PAYING for the billboard space, not who is selling it
- Agency (advertising agency, may be blank)
- Brand/Campaign (the advertised brand or campaign name)
  **IMPORTANT:** Use intelligent inference if brand/campaign is not explicitly stated:
  - If client is "Gucci LLC" → brand is likely "Gucci"
  - If client is "Emaar Properties PJSC" → brand is likely "Emaar"
  - If client is "Dubai Properties Development L.L.C" → brand is likely "Dubai Properties"
  - Extract the core brand name from the client company name by removing corporate suffixes like LLC, PJSC, L.L.C, Inc, Ltd, etc.
  - Only use the full client name as brand if there's truly no brand information anywhere in the document
- Category (the client's main industry/sector)
  **IMPORTANT:** Category represents the CLIENT'S industry, not the campaign type:
  - If client is "Emaar Properties PJSC" → category is "Real Estate"
  - If client is "Nestlé Middle East" → category is "FMCG" (Fast-Moving Consumer Goods)
  - If client is "Mercedes-Benz UAE" → category is "Automotive"
  - If client is "Emirates NBD" → category is "Banking/Finance"
  - Common categories: Real Estate, FMCG, Automotive, Banking/Finance, Hospitality, Retail, Healthcare, Technology, Entertainment
  - Infer from the client company name if not explicitly stated in the BO

**Location/Asset Details (usually in a table):**
For EACH billboard location, extract:
- Location code: Extract ONLY the alphanumeric code (e.g., "04B", "03A", "15C")
  - Strip away descriptive names like "Sheikh Zayed Road Tower" or "Al Barsha Mall"
  - Just the code: 2 digits + optional letter (04B, 03A, 15C, etc.)
- Start date (campaign start date)
- End date (campaign end date)
- Campaign duration **IMPORTANT:** Calculate and format intelligently:
  - Calculate number of days between start and end date
  - **Format rules (approximate to nearest unit):**
    * 28-31 days → "1 month"
    * 56-62 days → "2 months"
    * 84-93 days → "3 months"
    * 14-15 days → "2 weeks"
    * 21-22 days → "3 weeks"
    * 7-8 days → "1 week"
    * Only use "X days" if it doesn't fit these approximations
  - Examples: 30 days = "1 month", 15 days = "2 weeks", 60 days = "2 months", 10 days = "10 days"
  - If explicitly stated in BO (rare), use that value; otherwise calculate from dates
- Net amount (rental cost for THIS location - may need to split bundled payments)
- Production/Upload cost (if specified per-location)
- Type (digital vs static - if mentioned)

**CRITICAL: Understanding the Data Structure**

Booking orders have TWO types of costs:

1. **Per-Location Rental Amounts** (in locations array):
   - Each location has its own rental amount (e.g., 04B: AED 80,000, 03A: AED 80,000)
   - This is ONLY the rental cost for that specific billboard
   - Extract as `net_amount` for each location

2. **GLOBAL Fees** (top-level fields, NOT per-location):
   - **municipality_fee**: Dubai Municipality regulatory fee for the ENTIRE booking
     - Look for: "DM Fee", "Municipality Fee", "Dubai Municipality", "Net DM fee"
     - Typically a small amount (hundreds to low thousands)
     - ONE total for all locations combined
     - **If document shows per-location DM fees:** ADD them all up into ONE total
     - Example: 04B: AED 200, 03A: AED 200, 15C: AED 120 → municipality_fee: 520

   - **production_upload_fee**: Total production/upload cost for ALL locations
     - Look for: "Production Fee", "Upload Fee", "Production Cost", "Net Production fee"
     - This is production fee (for static) + upload fee (for digital) combined into ONE total
     - ONE total for all locations combined
     - **If document shows per-location fees:** ADD them all up into ONE total
     - Example: 04B upload: AED 500, 03A upload: AED 500, 15C production: AED 1000 → production_upload_fee: 2000

**CRITICAL:** Each location only stores its rental amount. ALL fees are GLOBAL single totals, NOT per-location values.
Even if the source document lists fees per location, you MUST sum them into single global totals.

**The Math:**
- Sum of all location rental amounts (e.g., 80,000 + 80,000 + 120,000 = 280,000)
- + production_upload_fee (e.g., 2,000)
- + municipality_fee (e.g., 520)
- = net_pre_vat (e.g., 282,520)

**Financial Totals:**
- Net amount (Rental + Production + Municipality, before SLA and VAT)
- VAT amount (5% tax, usually on Rental + Production, not Municipality)
- Gross amount (total including VAT)
- SLA % (deduction percentage, e.g., 10% = 0.10)

**Additional Information:**
- Payment terms (e.g., "60 days PDC", "30 days credit") - **CHECK USER MESSAGE** as this is often provided there
- Salesperson name - **CHECK USER MESSAGE** as user may specify who the sales person is
- Commission %

**CRITICAL NOTES:**
- If you see "Municipality Fee" or "Upload Fee" in the locations table, extract it separately as a fee, NOT as a location
- When multiple locations share one payment, split the amount across them intelligently
- Think about whether each line item is a location rental or a fee
- The structure may be messy - take your time to understand it
- Extract what you see, but understand the business logic to validate your extraction
- **REMEMBER:** For Viola, extract ONLY location codes (04B, 03A, etc.) - no descriptive names!"""

    def _extract_json_from_response(self, text: str) -> Dict[str, Any]:
        """Extract JSON from LLM response (may have markdown code blocks)"""
        # Remove markdown code blocks if present
        text = text.strip()
        if text.startswith("```json"):
            text = text[7:]
        if text.startswith("```"):
            text = text[3:]
        if text.endswith("```"):
            text = text[:-3]
        text = text.strip()

        try:
            return json.loads(text)
        except json.JSONDecodeError as e:
            logger.error(f"[BOOKING PARSER] Failed to parse JSON: {e}")
            logger.error(f"[BOOKING PARSER] Response text: {text[:500]}")
            raise ValueError(f"Invalid JSON in response: {e}")

    def _post_process_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Post-process parsed data: normalize, calculate derived fields"""
        # Calculate derived fields
        net = data.get("net_pre_vat")
        vat = data.get("vat_value")
        gross = data.get("gross_amount")
        sla_pct = data.get("sla_pct", 0) or 0

        if net is not None:
            # Calculate VAT if missing (5% standard in UAE)
            if vat is None:
                data["vat_calc"] = round(net * 0.05, 2)
            else:
                data["vat_calc"] = vat

            # Calculate gross if missing
            if gross is None:
                data["gross_calc"] = round(net + data["vat_calc"], 2)
            else:
                data["gross_calc"] = gross

            # Calculate SLA deduction on RENTAL AMOUNTS ONLY (excluding production, DM, upload fees)
            # Sum all location net_amounts (this is the rental total)
            locations = data.get("locations", [])
            rental_total = sum(loc.get("net_amount", 0) for loc in locations)

            # Calculate SLA deduction on rental total only
            data["sla_deduction"] = round(rental_total * sla_pct, 2)

            # Net excl SLA = full net_pre_vat minus the SLA deduction
            data["net_excl_sla_calc"] = round(net - data["sla_deduction"], 2)

            # Calculate per-location post-SLA amounts
            # Each location's post-SLA amount = location.net_amount * (1 - sla_pct)
            for location in locations:
                location_rental = location.get("net_amount", 0)
                location["post_sla_amount"] = round(location_rental * (1 - sla_pct), 2)

        # Normalize dates
        # Normalize bo_date
        if data.get("bo_date"):
            data["bo_date"] = self._normalize_date(data["bo_date"])

        # Normalize location dates
        for location in data.get("locations", []):
            for date_field in ["start_date", "end_date"]:
                if location.get(date_field):
                    location[date_field] = self._normalize_date(location[date_field])

        return data

    def _normalize_date(self, date_str: str) -> str:
        """
        Try to normalize date to English format: "Xth Month YYYY"
        Examples: "1st January 2025", "23rd March 2025"
        """
        if not date_str:
            return None

        # Try common formats
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"]:
            try:
                dt = datetime.strptime(date_str, fmt)
                # Format as "Xth Month YYYY"
                return self._format_date_english(dt)
            except:
                continue

        # Return original if can't parse
        return date_str

    def _format_date_english(self, dt: datetime) -> str:
        """
        Format datetime object as "Xth Month YYYY"
        Examples: "1st January 2025", "23rd March 2025"
        """
        day = dt.day
        month = dt.strftime("%B")  # Full month name
        year = dt.year

        # Add ordinal suffix (st, nd, rd, th)
        if 10 <= day % 100 <= 20:
            suffix = "th"
        else:
            suffix = {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")

        return f"{day}{suffix} {month} {year}"

    def _generate_warnings(self, data: Dict[str, Any]) -> List[str]:
        """Generate warnings about inconsistencies"""
        warnings = []

        # Check totals consistency
        net = data.get("net_pre_vat")
        vat_calc = data.get("vat_calc")
        gross_calc = data.get("gross_calc")
        gross_stated = data.get("gross_amount")

        if gross_stated and gross_calc and abs(gross_stated - gross_calc) > 0.01:
            warnings.append(f"Gross amount mismatch: stated {gross_stated} vs calculated {gross_calc}")

        # Check location totals sum (accounting for fees)
        if net and data.get("locations"):
            location_total = sum(loc.get("net_amount", 0) for loc in data["locations"])
            municipality_fee = data.get("municipality_fee", 0) or 0
            production_upload_fee = data.get("production_upload_fee", 0) or 0

            # Expected net = location rentals + all fees
            expected_net = location_total + municipality_fee + production_upload_fee

            # Only warn if there's a significant mismatch after accounting for fees
            if abs(expected_net - net) > 0.01:
                warnings.append(f"Location totals ({location_total}) + fees (municipality: {municipality_fee}, prod/upload: {production_upload_fee}) = {expected_net} doesn't match global net ({net})")

        return warnings

    def _check_missing_required(self, data: Dict[str, Any]) -> List[str]:
        """Check for missing required fields"""
        missing = []

        # Check global required
        for field in self.required_global:
            if not data.get(field):
                missing.append(f"Global field: {field}")

        # Check per-location required
        for i, location in enumerate(data.get("locations", [])):
            for field in self.required_per_location:
                if not location.get(field):
                    missing.append(f"Location {i+1}: {field}")

        return missing

    def format_for_slack(self, data: Dict[str, Any], bo_ref: str) -> str:
        """Format booking order data for Slack display"""
        lines = [
            f"**Booking Order:** {bo_ref}",
            f"**Client:** {data.get('client', 'N/A')}",
            f"**Campaign:** {data.get('brand_campaign', 'N/A')}",
            f"**Net (pre-VAT):** AED {data.get('net_pre_vat', 0):,.2f}",
            f"**VAT (5%):** AED {data.get('vat_calc', 0):,.2f}",
            f"**Gross Total:** AED {data.get('gross_calc', 0):,.2f}",
        ]

        if data.get("sla_pct"):
            lines.append(f"**SLA Deduction:** AED {data.get('sla_deduction', 0):,.2f} ({data['sla_pct']*100:.1f}%)")
            lines.append(f"**Net (excl SLA):** AED {data.get('net_excl_sla_calc', 0):,.2f}")

        if data.get("locations"):
            lines.append(f"\n**Locations:** {len(data['locations'])}")
            for loc in data["locations"][:5]:  # Show first 5
                lines.append(f"  • {loc.get('name', 'Unknown')}: {loc.get('start_date', '?')} to {loc.get('end_date', '?')}")

        return "\n".join(lines)

    async def generate_excel(self, data: Dict[str, Any], bo_ref: str) -> Path:
        """
        Generate branded Excel using company-specific template

        Templates:
        - Backlite: bo_templates/backlite_bo_template.xlsx
        - Viola: bo_templates/viola_bo_template.xlsx
        """
        import openpyxl

        # Select template based on company
        template_path = TEMPLATE_BACKLITE if self.company.lower() == "backlite" else TEMPLATE_VIOLA

        if not template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")

        logger.info(f"[BOOKING PARSER] Using template: {template_path}")

        # Load template
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Helper function to convert lists to comma-separated strings
        def format_value(value):
            if isinstance(value, list):
                return ", ".join(str(v) for v in value)
            return value if value is not None else ""

        # Helper to group locations by start date
        def get_start_dates():
            if not data.get("locations"):
                return ""

            # Group locations by start_date
            date_groups = {}
            for loc in data["locations"]:
                start_date = loc.get("start_date", "")
                if start_date:
                    loc_name = loc.get("name", "").replace("UAE", "UAE")  # Ensure consistent formatting
                    if start_date not in date_groups:
                        date_groups[start_date] = []
                    date_groups[start_date].append(loc_name)

            # Format as "UAE02 & UAE03: date\nUAE21: date"
            lines = []
            for date, locations in date_groups.items():
                loc_str = " & ".join(locations)
                lines.append(f"{loc_str}: {date}")

            return "\n".join(lines)

        # Helper to group locations by end date
        def get_end_dates():
            if not data.get("locations"):
                return ""

            # Group locations by end_date
            date_groups = {}
            for loc in data["locations"]:
                end_date = loc.get("end_date", "")
                if end_date:
                    loc_name = loc.get("name", "").replace("UAE", "UAE")  # Ensure consistent formatting
                    if end_date not in date_groups:
                        date_groups[end_date] = []
                    date_groups[end_date].append(loc_name)

            # Format as "UAE02 & UAE03: date\nUAE21: date"
            lines = []
            for date, locations in date_groups.items():
                loc_str = " & ".join(locations)
                lines.append(f"{loc_str}: {date}")

            return "\n".join(lines)

        # Helper to group locations by campaign duration
        def get_durations():
            if not data.get("locations"):
                return ""

            # Group locations by campaign_duration
            duration_groups = {}
            for loc in data["locations"]:
                duration = loc.get("campaign_duration", "")
                if duration:
                    loc_name = loc.get("name", "").replace("UAE", "UAE")  # Ensure consistent formatting
                    if duration not in duration_groups:
                        duration_groups[duration] = []
                    duration_groups[duration].append(loc_name)

            # Format as "UAE02 & UAE03: duration\nUAE21: duration"
            lines = []
            for duration, locations in duration_groups.items():
                loc_str = " & ".join(locations)
                lines.append(f"{loc_str}: {duration}")

            return "\n".join(lines)

        # Helper to get production/upload fee (global, not per-location)
        def get_production_upload_fee():
            # Use global production_upload_fee (already combined)
            return data.get("production_upload_fee", 0) or 0

        # Calculate Net Rentals excl SLA (for the merged cell A-E33)
        # This is the net amount AFTER SLA deduction
        net_rentals_excl_sla = data.get("net_excl_sla_calc") or data.get("net_pre_vat") or 0

        # Inject values into template cells
        # Left column (B)
        ws["B11"] = format_value(data.get("agency"))                    # Agency
        ws["B13"] = format_value(data.get("client"))                    # Client

        # Brand/Campaign with dynamic row height
        brand_campaign_value = format_value(data.get("brand_campaign"))
        ws["B15"] = brand_campaign_value
        ws["B15"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        num_lines_b15 = brand_campaign_value.count('\n') + 1 if brand_campaign_value else 1
        initial_height_b15 = max(20, num_lines_b15 * 20)
        ws.row_dimensions[15].height = initial_height_b15
        logger.info(f"[EXCEL] Row 15 B15 (brand/campaign): {num_lines_b15} lines, height set to {initial_height_b15}")

        # Start dates with dynamic row height
        start_dates_value = get_start_dates()
        ws["B17"] = start_dates_value
        ws["B17"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        # Adjust row height based on number of lines (each line ~15 points)
        num_lines_b17 = start_dates_value.count('\n') + 1 if start_dates_value else 1
        initial_height_b17 = max(15, num_lines_b17 * 15)
        ws.row_dimensions[17].height = initial_height_b17
        logger.info(f"[EXCEL] Row 17 B17 (start dates): {num_lines_b17} lines, height set to {initial_height_b17}")

        # Durations with dynamic row height (count & and commas for multiple locations)
        durations_value = get_durations()
        ws["B19"] = durations_value
        ws["B19"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        # Count & and commas to estimate lines (locations grouped by duration)
        num_separators_b19 = durations_value.count('&') + durations_value.count(',') if durations_value else 0
        num_lines_b19 = max(durations_value.count('\n') + 1 if durations_value else 1, (num_separators_b19 // 2) + 1)
        initial_height_b19 = max(20, num_lines_b19 * 20)
        ws.row_dimensions[19].height = initial_height_b19
        logger.info(f"[EXCEL] Row 19 B19 (durations): {num_lines_b19} lines ({num_separators_b19} separators), height set to {initial_height_b19}")

        ws["B21"] = data.get("gross_calc", 0)                           # Gross (net + vat)
        ws["B23"] = get_production_upload_fee()                         # Production/Upload Cost(s)
        ws["B25"] = data.get("vat_calc", 0)                             # VAT
        ws["B27"] = data.get("net_pre_vat", 0)                          # Net excl VAT (Net amount)

        # Right column (E)
        ws["E11"] = format_value(data.get("bo_number"))                 # BO No.
        ws["E13"] = format_value(data.get("bo_date"))                   # BO date

        # Asset(s) with dynamic row height (count commas for multiple assets)
        asset_value = format_value(data.get("asset"))
        ws["E15"] = asset_value
        ws["E15"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        # Count commas but divide by 2 (assets are often short, don't need full line each)
        num_commas_e15 = asset_value.count(',') if asset_value else 0
        # Use newlines primarily, but add some buffer for commas (every 2-3 commas = 1 extra line)
        num_lines_e15 = max(asset_value.count('\n') + 1 if asset_value else 1, 1 + (num_commas_e15 // 3))
        # Use the max of B15 and E15 line counts for row 15
        before_height_15 = ws.row_dimensions[15].height
        final_height_15 = max(before_height_15, num_lines_e15 * 20)
        ws.row_dimensions[15].height = final_height_15
        logger.info(f"[EXCEL] Row 15 E15 (asset): {num_lines_e15} lines ({num_commas_e15} commas), before={before_height_15}, after={final_height_15}")

        # End dates with dynamic row height
        end_dates_value = get_end_dates()
        ws["E17"] = end_dates_value
        ws["E17"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        num_lines_e17 = end_dates_value.count('\n') + 1 if end_dates_value else 1
        # Use the max of B17 and E17 line counts for row 17
        before_height_17 = ws.row_dimensions[17].height
        final_height_17 = max(before_height_17, num_lines_e17 * 15)
        ws.row_dimensions[17].height = final_height_17
        logger.info(f"[EXCEL] Row 17 E17 (end dates): {num_lines_e17} lines, before={before_height_17}, after={final_height_17}")

        # Category in E19 - also set wrap text and check if it needs more height
        category_value = format_value(data.get("category"))
        ws["E19"] = category_value
        ws["E19"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        num_lines_e19 = category_value.count('\n') + 1 if category_value else 1
        # Use the max of B19 and E19 line counts for row 19
        before_height_19 = ws.row_dimensions[19].height
        final_height_19 = max(before_height_19, num_lines_e19 * 20)
        ws.row_dimensions[19].height = final_height_19
        logger.info(f"[EXCEL] Row 19 E19 (category): {num_lines_e19} lines, before={before_height_19}, after={final_height_19}")

        # SLA - show the deduction amount, not the percentage
        # E.g., if 10% of 462,520 = 46,252, show 46,252
        sla_deduction = data.get("sla_deduction", 0) or 0
        ws["E21"] = sla_deduction
        ws["E23"] = data.get("municipality_fee", 0)                     # DM (Dubai Municipality)

        # Payment terms with dynamic row height
        payment_terms_value = format_value(data.get("payment_terms"))
        ws["E25"] = payment_terms_value
        ws["E25"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        num_lines_e25 = payment_terms_value.count('\n') + 1 if payment_terms_value else 1
        ws.row_dimensions[25].height = max(15, num_lines_e25 * 15)

        ws["E27"] = format_value(data.get("sales_person"))              # Sales Person Name
        ws["E29"] = data.get("commission_pct", 0)                       # Commission%

        # Head of Sales Signature section
        ws["A39"] = "Head of Sales Signature"

        # Add HoS signature if provided (in italics)
        hos_signature = data.get("hos_signature")
        if hos_signature:
            ws["B39"] = hos_signature
            ws["B39"].font = openpyxl.styles.Font(italic=True)

        # Net rentals excl SLA in merged cell (A-E 32-37 range)
        # Show breakdown with per-location post-SLA amounts if SLA was applied
        sla_pct = data.get("sla_pct", 0) or 0

        # Build the Net excl SLA display text
        net_excl_sla_text = f"Net excl SLA: {net_rentals_excl_sla:,.2f}"

        # If SLA was applied, show per-location breakdown
        if sla_pct > 0:
            locations = data.get("locations", [])
            location_lines = []

            for location in locations:
                loc_name = location.get("location", "Unknown")
                pre_sla = location.get("net_amount", 0)
                post_sla = location.get("post_sla_amount", pre_sla)

                # Only show location breakdown if post-SLA differs from pre-SLA
                if abs(post_sla - pre_sla) > 0.01:
                    location_lines.append(f"  {loc_name}: {post_sla:,.2f}")

            # Add location breakdown if any locations had SLA applied
            if location_lines:
                net_excl_sla_text += "\n" + "\n".join(location_lines)

        # For merged cells, we must write to the top-left cell of the range
        for merged_range in ws.merged_cells.ranges:
            if "A33" in merged_range:
                # Get the top-left cell of the merged range
                min_col, min_row = merged_range.bounds[0], merged_range.bounds[1]
                top_left_cell = ws.cell(row=min_row, column=min_col)
                top_left_cell.value = net_excl_sla_text
                top_left_cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                break

        # Save to temporary file
        temp_excel = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_excel_path = Path(temp_excel.name)
        wb.save(temp_excel_path)
        logger.info(f"[BOOKING PARSER] Generated branded Excel: {temp_excel_path}")

        return temp_excel_path

    async def generate_combined_pdf(self, data: Dict[str, Any], bo_ref: str, original_bo_path: Path) -> Path:
        """
        Generate combined PDF: Excel (converted to PDF) + Original BO PDF concatenated

        Args:
            data: Parsed booking order data
            bo_ref: BO reference for filename
            original_bo_path: Path to the original BO file (PDF or will be converted)

        Returns:
            Path to the combined PDF file
        """
        logger.info(f"[BOOKING PARSER] Generating combined PDF for {bo_ref}, original_bo_path: {original_bo_path} (type: {type(original_bo_path)})")

        # Step 1: Generate Excel
        logger.info(f"[BOOKING PARSER] Step 1: Generating Excel for {bo_ref}")
        excel_path = await self.generate_excel(data, bo_ref)
        logger.info(f"[BOOKING PARSER] Excel generated: {excel_path}")

        # Step 2: Convert Excel to PDF using LibreOffice
        logger.info(f"[BOOKING PARSER] Step 2: Converting Excel to PDF")
        excel_pdf_path = await self._convert_excel_to_pdf(excel_path)
        logger.info(f"[BOOKING PARSER] Excel PDF created: {excel_pdf_path}")

        # Step 3: Ensure original BO is PDF (convert if needed)
        logger.info(f"[BOOKING PARSER] Step 3: Ensuring original BO is PDF, path: {original_bo_path}")
        original_pdf_path = await self._ensure_pdf(original_bo_path)
        logger.info(f"[BOOKING PARSER] Original BO PDF ready: {original_pdf_path}")

        # Step 4: Concatenate PDFs (Excel PDF first, then original BO)
        combined_pdf_path = COMBINED_BOS_DIR / f"{bo_ref}_combined.pdf"
        await self._concatenate_pdfs([excel_pdf_path, original_pdf_path], combined_pdf_path)

        # Clean up temporary files
        try:
            excel_path.unlink()
            if excel_pdf_path != excel_path:
                excel_pdf_path.unlink()
            if original_pdf_path != original_bo_path:
                original_pdf_path.unlink()
        except Exception as e:
            logger.warning(f"[BOOKING PARSER] Failed to clean up temp files: {e}")

        logger.info(f"[BOOKING PARSER] Combined PDF generated: {combined_pdf_path}")
        return combined_pdf_path

    async def _convert_excel_to_pdf(self, excel_path: Path) -> Path:
        """Convert Excel file to PDF using LibreOffice"""
        logger.info(f"[BOOKING PARSER] Converting Excel to PDF: {excel_path}")

        output_dir = excel_path.parent
        pdf_path = excel_path.with_suffix('.pdf')

        try:
            # Use LibreOffice to convert Excel to PDF
            result = subprocess.run([
                'soffice',
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', str(output_dir),
                str(excel_path)
            ], capture_output=True, text=True, timeout=30)

            if result.returncode != 0:
                raise Exception(f"LibreOffice conversion failed: {result.stderr}")

            if not pdf_path.exists():
                raise Exception(f"PDF not created at expected path: {pdf_path}")

            logger.info(f"[BOOKING PARSER] Excel converted to PDF: {pdf_path}")
            return pdf_path

        except Exception as e:
            logger.error(f"[BOOKING PARSER] Failed to convert Excel to PDF: {e}")
            raise

    async def _ensure_pdf(self, file_path: Path) -> Path:
        """Ensure file is PDF, convert if needed"""
        if file_path.suffix.lower() == '.pdf':
            return file_path

        logger.info(f"[BOOKING PARSER] Converting {file_path.suffix} to PDF")

        output_dir = file_path.parent
        pdf_path = file_path.with_suffix('.pdf')

        try:
            result = subprocess.run([
                'soffice',
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', str(output_dir),
                str(file_path)
            ], capture_output=True, text=True, timeout=30)

            if result.returncode != 0:
                raise Exception(f"LibreOffice conversion failed: {result.stderr}")

            if not pdf_path.exists():
                raise Exception(f"PDF not created at expected path: {pdf_path}")

            logger.info(f"[BOOKING PARSER] File converted to PDF: {pdf_path}")
            return pdf_path

        except Exception as e:
            logger.error(f"[BOOKING PARSER] Failed to convert to PDF: {e}")
            raise

    async def _concatenate_pdfs(self, pdf_paths: List[Path], output_path: Path) -> None:
        """Concatenate multiple PDFs into one"""
        logger.info(f"[BOOKING PARSER] Concatenating {len(pdf_paths)} PDFs")

        try:
            merger = PdfMerger()

            for pdf_path in pdf_paths:
                logger.info(f"[BOOKING PARSER] Adding PDF: {pdf_path}")
                merger.append(str(pdf_path))

            merger.write(str(output_path))
            merger.close()

            logger.info(f"[BOOKING PARSER] PDFs concatenated successfully: {output_path}")

        except Exception as e:
            logger.error(f"[BOOKING PARSER] Failed to concatenate PDFs: {e}")
            raise
