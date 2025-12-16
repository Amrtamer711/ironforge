"""
Booking Order Parser for MMG Backlite and Viola
Handles classification and parsing of booking order documents using OpenAI Responses API
"""

import json
import logging
import os
import subprocess
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from pypdf import PdfMerger

import config
from integrations.llm import (
    ContentPart,
    LLMClient,
    LLMMessage,
    ReasoningEffort,
)
from integrations.llm.prompts.bo_parsing import (
    get_backlite_parsing_prompt,
    get_classification_prompt,
    get_data_extractor_prompt,
    get_viola_parsing_prompt,
)
from integrations.llm.schemas.bo_parsing import (
    get_booking_order_extraction_schema,
    get_classification_schema,
)

logger = logging.getLogger("proposal-bot")

# File storage directories
BOOKING_ORDERS_BASE = Path("/data/booking_orders") if os.path.exists("/data") else Path(__file__).parent / "booking_orders"
# Combined PDFs directory (stores Excel + Original BO concatenated)
COMBINED_BOS_DIR = BOOKING_ORDERS_BASE / "combined_bos"
# Original uploaded BOs directory (permanent storage for user uploads during workflow)
ORIGINAL_BOS_DIR = BOOKING_ORDERS_BASE / "original_uploads"

# BO Template files (for future use - not currently used)
# These are the actual branded templates for Backlite and Viola
# Currently using simple Excel generation, will switch to template-based when ready
TEMPLATES_DIR = Path(__file__).parent / "bo_templates"
TEMPLATE_BACKLITE = TEMPLATES_DIR / "backlite_bo_template.xlsx"
TEMPLATE_VIOLA = TEMPLATES_DIR / "viola_bo_template.xlsx"

# Ensure directories exist
COMBINED_BOS_DIR.mkdir(parents=True, exist_ok=True)
ORIGINAL_BOS_DIR.mkdir(parents=True, exist_ok=True)


def sanitize_filename(filename: str) -> str:
    """
    Sanitize a string to be safe for use as a filename.
    Removes/replaces invalid characters for filesystems.

    Args:
        filename: The filename to sanitize

    Returns:
        Sanitized filename safe for filesystem use
    """
    import re
    # Remove or replace invalid characters: / \ : * ? " < > | and control characters
    sanitized = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', str(filename))
    # Replace multiple underscores/spaces with single underscore
    sanitized = re.sub(r'[_\s]+', '_', sanitized)
    # Remove leading/trailing underscores, dots, or spaces
    sanitized = sanitized.strip('_. ')
    # Limit length to 200 characters to be safe
    sanitized = sanitized[:200]
    # If empty after sanitization, use a default
    return sanitized if sanitized else 'booking_order'


@dataclass
class ParseResult:
    """Result from parsing a booking order"""
    data: dict[str, Any]
    warnings: list[str]
    missing_required: list[str]
    needs_review: bool


class BookingOrderParser:
    """Parser for booking order documents with company-specific rules"""

    def __init__(self, company: str = "backlite"):
        self.company = company
        self.required_global = ["client", "net_pre_vat"]
        self.required_per_location = ["name", "start_date", "end_date", "net_amount"]

    def detect_file_type(self, file_path: Path) -> str:
        """Detect file type from extension"""
        suffix = file_path.suffix.lower()
        if suffix in [".xlsx", ".xls"]:
            return "excel"
        elif suffix == ".pdf":
            return "pdf"
        elif suffix in [".jpg", ".jpeg", ".png", ".gif", ".bmp"]:
            return "image"
        else:
            return "unknown"

    async def classify_document(self, file_path: Path, user_message: str = "", user_id: str = None) -> dict[str, str]:
        """
        Classify document as BOOKING_ORDER or ARTWORK using OpenAI Responses API.
        Biased toward ARTWORK unless clear booking order fields present.

        Accepts any file type - will convert to PDF if needed for classification.

        Args:
            file_path: Path to the document file (any format)
            user_message: Optional user's message/prompt that accompanied the file upload
        """
        logger.info(f"[BOOKING PARSER] Classifying document: {file_path}")
        if user_message:
            logger.info(f"[BOOKING PARSER] User message context: {user_message}")

        # Convert to PDF if needed (OpenAI Responses API only accepts PDFs)
        pdf_path = file_path
        cleanup_pdf = False

        # Check file extension
        suffix = file_path.suffix.lower()

        # Convert images to PDF
        if suffix in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']:
            logger.info("[BOOKING PARSER] Converting image to PDF for classification...")
            try:
                from PIL import Image
                pdf_path = file_path.with_suffix('.pdf')
                img = Image.open(file_path)
                # Convert to RGB if needed (PNG with transparency, etc.)
                if img.mode in ('RGBA', 'LA', 'P'):
                    img = img.convert('RGB')
                img.save(pdf_path, 'PDF', resolution=100.0)
                cleanup_pdf = True
                logger.info(f"[BOOKING PARSER] Converted image to PDF: {pdf_path}")
            except Exception as e:
                logger.error(f"[BOOKING PARSER] Failed to convert image to PDF: {e}")
                return {"classification": "ARTWORK", "confidence": "high", "reasoning": "Image file detected"}

        # Convert Excel to PDF
        elif suffix in ['.xlsx', '.xls']:
            logger.info("[BOOKING PARSER] Converting Excel to PDF for classification...")
            try:
                pdf_path = await self._convert_excel_to_pdf(file_path)
                cleanup_pdf = True
                logger.info(f"[BOOKING PARSER] Converted Excel to PDF: {pdf_path}")
            except Exception as e:
                logger.error(f"[BOOKING PARSER] Failed to convert Excel to PDF: {e}")
                return {"classification": "UNKNOWN", "confidence": "low", "reasoning": f"Excel conversion failed: {e}"}

        # Use LLM client for classification
        llm_client = LLMClient.from_config()
        file_ref = None

        try:
            # Upload PDF
            file_ref = await llm_client.upload_file(str(pdf_path))
            logger.info(f"[BOOKING PARSER] Uploaded file: {file_ref.file_id}")
        except Exception as e:
            logger.error(f"[BOOKING PARSER] Failed to upload file: {e}")
            if cleanup_pdf and pdf_path.exists():
                pdf_path.unlink(missing_ok=True)
            return {"classification": "UNKNOWN", "confidence": "low", "reasoning": str(e)}
        finally:
            # Clean up temporary PDF if we created one
            if cleanup_pdf and pdf_path.exists() and pdf_path != file_path:
                try:
                    pdf_path.unlink(missing_ok=True)
                except Exception as e:
                    logger.warning(f"[BOOKING PARSER] Failed to clean up temp PDF: {e}")

        # Get the classification prompt
        classification_prompt = get_classification_prompt(user_message)

        try:
            # Use LLM client with structured JSON output
            response = await llm_client.complete(
                messages=[
                    LLMMessage.system("You are a document classifier. Analyze the file and provide classification in JSON format."),
                    LLMMessage.user([
                        ContentPart.file(file_ref.file_id),
                        ContentPart.text(classification_prompt)
                    ])
                ],
                json_schema=get_classification_schema(),
                store=config.IS_DEVELOPMENT,  # Store in OpenAI only in dev mode
                # Prompt caching: classification system prompt is static
                cache_key="bo-classify",
                cache_retention="24h",
                # Cost tracking
                call_type="classification",
                workflow="bo_parsing",
                user_id=user_id,
                context=f"File: {file_path.name}",
                metadata={"file_type": suffix, "has_user_message": bool(user_message)}
            )

            if not response.content:
                logger.warning("[BOOKING PARSER] Empty classification response")
                return {"classification": "ARTWORK", "confidence": "low", "reasoning": "No response from model"}

            # Parse JSON response from structured output
            logger.info(f"[BOOKING PARSER] Classification response: {response.content}")

            # Parse JSON directly (structured output guarantees valid JSON)
            result = json.loads(response.content)

            # Ensure company is set for BOOKING_ORDER (fallback to backlite if null)
            if result["classification"] == "BOOKING_ORDER" and not result.get("company"):
                result["company"] = "backlite"

            logger.info(f"[BOOKING PARSER] Parsed classification: {result}")
            return result

        except Exception as e:
            logger.error(f"[BOOKING PARSER] Classification error: {e}", exc_info=True)
            return {"classification": "ARTWORK", "confidence": "low", "reasoning": f"Error: {str(e)}"}
        finally:
            # Cleanup uploaded file
            if file_ref:
                await llm_client.delete_file(file_ref)

    async def parse_file(self, file_path: Path, file_type: str, user_message: str = "", user_id: str = None) -> ParseResult:
        """
        Parse booking order file using LLM with structured JSON output.
        No hallucinations - only extract what's clearly present.

        Args:
            file_path: Path to the booking order file
            file_type: File type (pdf, image, excel)
            user_message: Optional user's message that accompanied the file upload
        """
        logger.info(f"[BOOKING PARSER] Parsing {file_type} file: {file_path}")
        if user_message:
            logger.info(f"[BOOKING PARSER] User message context: {user_message}")

        # Use LLM client for parsing
        llm_client = LLMClient.from_config()
        file_ref = None

        try:
            file_ref = await llm_client.upload_file(str(file_path))
            logger.info(f"[BOOKING PARSER] Uploaded file for parsing: {file_ref.file_id}")
        except Exception as e:
            logger.error(f"[BOOKING PARSER] Failed to upload file for parsing: {e}")
            raise

        # Parsing prompt for structured extraction (no schema in prompt)
        if self.company.lower() == "viola":
            parsing_prompt = get_viola_parsing_prompt()
        else:
            parsing_prompt = get_backlite_parsing_prompt()

        # Add user message context if provided
        if user_message:
            parsing_prompt = f"""{parsing_prompt}

**USER'S MESSAGE CONTEXT:**
The user provided this message with the file: "{user_message}"

**CRITICAL - PRIORITIZE USER MESSAGE:** The user message often contains important information that may be missing from or clarify the document. Pay special attention to:
- **SLA amount/percentage**: User may specify "10% SLA" or "SLA is 5%" - ALWAYS extract this if mentioned
- **Payment terms**: User may specify "60 days PDC" or "30 days credit" - extract if mentioned
- **Sales person name**: User may say "this is for John" or "sales person is Sarah" - extract if mentioned
- **Location payment splits**: User may clarify "UAE02 is 100k, UAE03 is 200k" - use these exact amounts
- **Fee amounts**: User may specify "production fee is 2000" or "municipality fee is 500" - prioritize these values
- **Client/company details**: If user mentions client name or category, use that information
- **Any clarifications**: User message helps resolve ambiguities in the document

**PRIORITIZATION RULE:** If the user message contains explicit values for any field, those values OVERRIDE what's in the document.
"""

        try:
            # Use LLM client with structured JSON output and high reasoning
            response = await llm_client.complete(
                messages=[
                    LLMMessage.system(get_data_extractor_prompt()),
                    LLMMessage.user([
                        ContentPart.file(file_ref.file_id),
                        ContentPart.text(parsing_prompt)
                    ])
                ],
                reasoning=ReasoningEffort.HIGH,
                json_schema=get_booking_order_extraction_schema(),
                store=config.IS_DEVELOPMENT,  # Store in OpenAI only in dev mode
                # Prompt caching: data extractor prompt is static
                cache_key="bo-parse",
                cache_retention="24h",
                # Cost tracking
                call_type="parsing",
                workflow="bo_parsing",
                user_id=user_id,
                context=f"File: {file_path.name}, Company: {self.company}",
                metadata={"file_type": file_type, "has_user_message": bool(user_message), "company": self.company}
            )

            if not response.content:
                raise ValueError("Empty parsing response from model")

            # Log response
            logger.info(f"[BOOKING PARSER] Parse response length: {len(response.content)} chars")
            logger.info(f"[BOOKING PARSER] Parse response text: {response.content[:500]}...")  # Log first 500 chars

            # Parse JSON (should be valid JSON from structured outputs)
            parsed_data = json.loads(response.content)

            # Post-process and validate
            processed = self._post_process_data(parsed_data)
            warnings = self._generate_warnings(processed)
            missing = self._check_missing_required(processed)
            needs_review = len(warnings) > 0 or len(missing) > 0

            return ParseResult(
                data=processed,
                warnings=warnings,
                missing_required=missing,
                needs_review=needs_review
            )

        except Exception as e:
            logger.error(f"[BOOKING PARSER] Parsing error: {e}", exc_info=True)
            raise
        finally:
            # Cleanup uploaded file
            if file_ref:
                await llm_client.delete_file(file_ref)


    def _extract_json_from_response(self, text: str) -> dict[str, Any]:
        """Extract JSON from LLM response (may have markdown code blocks)"""
        # Remove markdown code blocks if present
        text = text.strip()
        if text.startswith("```json"):
            text = text[7:]
        if text.startswith("```"):
            text = text[3:]
        if text.endswith("```"):
            text = text[:-3]
        text = text.strip()

        try:
            return json.loads(text)
        except json.JSONDecodeError as e:
            logger.error(f"[BOOKING PARSER] Failed to parse JSON: {e}")
            logger.error(f"[BOOKING PARSER] Response text: {text[:500]}")
            raise ValueError(f"Invalid JSON in response: {e}")

    def _post_process_data(self, data: dict[str, Any]) -> dict[str, Any]:
        """Post-process parsed data: normalize, calculate derived fields"""
        currency = data.get("currency")
        if isinstance(currency, str) and currency.strip():
            data["currency"] = currency.strip().upper()
        else:
            data["currency"] = config.DEFAULT_CURRENCY

        # Calculate derived fields
        net = data.get("net_pre_vat")
        vat = data.get("vat_value")
        gross = data.get("gross_amount")
        sla_pct = data.get("sla_pct", 0) or 0

        if net is not None:
            # Calculate VAT if missing (5% standard in UAE)
            if vat is None:
                data["vat_calc"] = round(net * 0.05, 2)
            else:
                data["vat_calc"] = vat

            # Calculate gross if missing
            if gross is None:
                data["gross_calc"] = round(net + data["vat_calc"], 2)
            else:
                data["gross_calc"] = gross

            # Calculate SLA deduction on RENTAL AMOUNTS ONLY (excluding production, DM, upload fees)
            # Sum all location net_amounts (this is the rental total)
            locations = data.get("locations", [])
            rental_total = sum(loc.get("net_amount", 0) for loc in locations)

            # Calculate SLA deduction on rental total only
            data["sla_deduction"] = round(rental_total * sla_pct, 2)

            # Net excl SLA = full net_pre_vat minus the SLA deduction
            data["net_excl_sla_calc"] = round(net - data["sla_deduction"], 2)

            # Calculate per-location post-SLA amounts
            # Each location's post-SLA amount = location.net_amount * (1 - sla_pct)
            for location in locations:
                location_rental = location.get("net_amount", 0)
                location["post_sla_amount"] = round(location_rental * (1 - sla_pct), 2)

        # Normalize dates
        # Normalize bo_date
        if data.get("bo_date"):
            data["bo_date"] = self._normalize_date(data["bo_date"])

        # Normalize location dates
        for location in data.get("locations", []):
            for date_field in ["start_date", "end_date"]:
                if location.get(date_field):
                    location[date_field] = self._normalize_date(location[date_field])

        return data

    def _normalize_date(self, date_str: str) -> str:
        """
        Try to normalize date to English format: "Xth Month YYYY"
        Examples: "1st January 2025", "23rd March 2025"
        """
        if not date_str:
            return None

        # Try common formats
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"]:
            try:
                dt = datetime.strptime(date_str, fmt)
                # Format as "Xth Month YYYY"
                return self._format_date_english(dt)
            except ValueError:
                continue  # Expected - try next format

        # Return original if can't parse
        return date_str

    def _format_date_english(self, dt: datetime) -> str:
        """
        Format datetime object as "Xth Month YYYY"
        Examples: "1st January 2025", "23rd March 2025"
        """
        day = dt.day
        month = dt.strftime("%B")  # Full month name
        year = dt.year

        # Add ordinal suffix (st, nd, rd, th)
        if 10 <= day % 100 <= 20:
            suffix = "th"
        else:
            suffix = {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")

        return f"{day}{suffix} {month} {year}"

    def _generate_warnings(self, data: dict[str, Any]) -> list[str]:
        """Generate warnings about inconsistencies"""
        warnings = []

        # Check totals consistency
        net = data.get("net_pre_vat")
        vat_calc = data.get("vat_calc")
        gross_calc = data.get("gross_calc")
        gross_stated = data.get("gross_amount")

        if gross_stated and gross_calc and abs(gross_stated - gross_calc) > 0.01:
            warnings.append(f"Gross amount mismatch: stated {gross_stated} vs calculated {gross_calc}")

        # Check location totals sum (accounting for fees)
        if net and data.get("locations"):
            location_total = sum(loc.get("net_amount", 0) for loc in data["locations"])
            municipality_fee = data.get("municipality_fee", 0) or 0
            production_upload_fee = data.get("production_upload_fee", 0) or 0

            # Expected net = location rentals + all fees
            expected_net = location_total + municipality_fee + production_upload_fee

            # Only warn if there's a significant mismatch after accounting for fees
            if abs(expected_net - net) > 0.01:
                warnings.append(f"Location totals ({location_total}) + fees (municipality: {municipality_fee}, prod/upload: {production_upload_fee}) = {expected_net} doesn't match global net ({net})")

        return warnings

    def _check_missing_required(self, data: dict[str, Any]) -> list[str]:
        """Check for missing required fields"""
        missing = []

        # Check global required
        for field in self.required_global:
            if not data.get(field):
                missing.append(f"Global field: {field}")

        # Check per-location required
        for i, location in enumerate(data.get("locations", [])):
            for field in self.required_per_location:
                if not location.get(field):
                    missing.append(f"Location {i+1}: {field}")

        return missing

    def format_for_slack(self, data: dict[str, Any], bo_ref: str) -> str:
        """Format booking order data for Slack display"""
        currency = data.get("currency", config.DEFAULT_CURRENCY)
        lines = [
            f"**Booking Order:** {bo_ref}",
            f"**Client:** {data.get('client', 'N/A')}",
            f"**Campaign:** {data.get('brand_campaign', 'N/A')}",
            f"**Net (pre-VAT):** {config.format_currency_value(data.get('net_pre_vat'), currency)}",
            f"**VAT (5%):** {config.format_currency_value(data.get('vat_calc'), currency)}",
            f"**Gross Total:** {config.format_currency_value(data.get('gross_calc'), currency)}",
        ]

        if data.get("sla_pct"):
            lines.append(
                f"**SLA Deduction:** {config.format_currency_value(data.get('sla_deduction'), currency)} ({data['sla_pct']*100:.1f}%)"
            )
            lines.append(
                f"**Net (excl SLA):** {config.format_currency_value(data.get('net_excl_sla_calc'), currency)}"
            )

        if data.get("locations"):
            lines.append(f"\n**Locations:** {len(data['locations'])}")
            for loc in data["locations"][:5]:  # Show first 5
                lines.append(f"  • {loc.get('name', 'Unknown')}: {loc.get('start_date', '?')} to {loc.get('end_date', '?')}")

        return "\n".join(lines)

    async def generate_excel(self, data: dict[str, Any], bo_ref: str) -> Path:
        """
        Generate branded Excel using company-specific template

        Templates:
        - Backlite: bo_templates/backlite_bo_template.xlsx
        - Viola: bo_templates/viola_bo_template.xlsx
        """
        import openpyxl

        # Select template based on company
        template_path = TEMPLATE_BACKLITE if self.company.lower() == "backlite" else TEMPLATE_VIOLA

        if not template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")

        logger.info(f"[BOOKING PARSER] Using template: {template_path}")

        # Load template
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Helper function to convert lists to comma-separated strings
        def format_value(value):
            if isinstance(value, list):
                return ", ".join(str(v) for v in value)
            return value if value is not None else ""

        # Helper to group locations by start date
        def get_start_dates():
            if not data.get("locations"):
                return ""

            # Group locations by start_date
            date_groups = {}
            for loc in data["locations"]:
                start_date = loc.get("start_date", "")
                if start_date:
                    loc_name = loc.get("name", "").replace("UAE", "UAE")  # Ensure consistent formatting
                    if start_date not in date_groups:
                        date_groups[start_date] = []
                    date_groups[start_date].append(loc_name)

            # Format as "UAE02 & UAE03: date\nUAE21: date"
            lines = []
            for date, locations in date_groups.items():
                loc_str = " & ".join(locations)
                lines.append(f"{loc_str}: {date}")

            return "\n".join(lines)

        # Helper to group locations by end date
        def get_end_dates():
            if not data.get("locations"):
                return ""

            # Group locations by end_date
            date_groups = {}
            for loc in data["locations"]:
                end_date = loc.get("end_date", "")
                if end_date:
                    loc_name = loc.get("name", "").replace("UAE", "UAE")  # Ensure consistent formatting
                    if end_date not in date_groups:
                        date_groups[end_date] = []
                    date_groups[end_date].append(loc_name)

            # Format as "UAE02 & UAE03: date\nUAE21: date"
            lines = []
            for date, locations in date_groups.items():
                loc_str = " & ".join(locations)
                lines.append(f"{loc_str}: {date}")

            return "\n".join(lines)

        # Helper to group locations by campaign duration
        def get_durations():
            if not data.get("locations"):
                return ""

            # Group locations by campaign_duration
            duration_groups = {}
            for loc in data["locations"]:
                duration = loc.get("campaign_duration", "")
                if duration:
                    loc_name = loc.get("name", "").replace("UAE", "UAE")  # Ensure consistent formatting
                    if duration not in duration_groups:
                        duration_groups[duration] = []
                    duration_groups[duration].append(loc_name)

            # Format as "UAE02 & UAE03: duration\nUAE21: duration"
            lines = []
            for duration, locations in duration_groups.items():
                loc_str = " & ".join(locations)
                lines.append(f"{loc_str}: {duration}")

            return "\n".join(lines)

        # Helper to get production/upload fee (global, not per-location)
        def get_production_upload_fee():
            # Use global production_upload_fee (already combined)
            return data.get("production_upload_fee", 0) or 0

        # Calculate Net Rentals excl SLA (for the merged cell A-E33)
        # This is the net amount AFTER SLA deduction
        net_rentals_excl_sla = data.get("net_excl_sla_calc") or data.get("net_pre_vat") or 0

        # Inject values into template cells
        # Left column (B)
        ws["B11"] = format_value(data.get("agency"))                    # Agency
        ws["B13"] = format_value(data.get("client"))                    # Client

        # Brand/Campaign with dynamic row height
        brand_campaign_value = format_value(data.get("brand_campaign"))
        ws["B15"] = brand_campaign_value
        ws["B15"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        num_lines_b15 = brand_campaign_value.count('\n') + 1 if brand_campaign_value else 1
        initial_height_b15 = max(20, num_lines_b15 * 20)
        ws.row_dimensions[15].height = initial_height_b15
        logger.info(f"[EXCEL] Row 15 B15 (brand/campaign): {num_lines_b15} lines, height set to {initial_height_b15}")

        # Start dates with dynamic row height
        start_dates_value = get_start_dates()
        ws["B17"] = start_dates_value
        ws["B17"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        # Adjust row height based on number of lines (each line ~15 points)
        num_lines_b17 = start_dates_value.count('\n') + 1 if start_dates_value else 1
        initial_height_b17 = max(15, num_lines_b17 * 15)
        ws.row_dimensions[17].height = initial_height_b17
        logger.info(f"[EXCEL] Row 17 B17 (start dates): {num_lines_b17} lines, height set to {initial_height_b17}")

        # Durations with dynamic row height (count & and commas for multiple locations)
        durations_value = get_durations()
        ws["B19"] = durations_value
        ws["B19"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        # Count & and commas to estimate lines (locations grouped by duration)
        num_separators_b19 = durations_value.count('&') + durations_value.count(',') if durations_value else 0
        num_lines_b19 = max(durations_value.count('\n') + 1 if durations_value else 1, (num_separators_b19 // 2) + 1)
        initial_height_b19 = max(20, num_lines_b19 * 20)
        ws.row_dimensions[19].height = initial_height_b19
        logger.info(f"[EXCEL] Row 19 B19 (durations): {num_lines_b19} lines ({num_separators_b19} separators), height set to {initial_height_b19}")

        # Get currency metadata for Excel formatting
        currency = data.get("currency", config.DEFAULT_CURRENCY)
        currency_meta = config.get_currency_metadata(currency)
        symbol = currency_meta.get("symbol", currency)
        position = currency_meta.get("position", "suffix")
        decimals = int(currency_meta.get("decimals", 2))

        # Build Excel number format based on currency position
        if position == "prefix":
            excel_number_format = f'"{symbol}"#,##0.{"0" * decimals}'
        else:
            excel_number_format = f'#,##0.{"0" * decimals}" {symbol}"'

        ws["B21"] = data.get("gross_calc", 0)                           # Gross (net + vat)
        ws["B21"].number_format = excel_number_format
        ws["B23"] = get_production_upload_fee()                         # Production/Upload Cost(s)
        ws["B23"].number_format = excel_number_format
        ws["B25"] = data.get("vat_calc", 0)                             # VAT
        ws["B25"].number_format = excel_number_format
        ws["B27"] = data.get("net_pre_vat", 0)                          # Net excl VAT (Net amount)
        ws["B27"].number_format = excel_number_format

        # Right column (E)
        ws["E11"] = format_value(data.get("bo_number"))                 # BO No.
        ws["E13"] = format_value(data.get("bo_date"))                   # BO date

        # Asset(s) with dynamic row height (count commas for multiple assets)
        asset_value = format_value(data.get("asset"))
        ws["E15"] = asset_value
        ws["E15"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        # Count commas but divide by 2 (assets are often short, don't need full line each)
        num_commas_e15 = asset_value.count(',') if asset_value else 0
        # Use newlines primarily, but add some buffer for commas (every 2-3 commas = 1 extra line)
        num_lines_e15 = max(asset_value.count('\n') + 1 if asset_value else 1, 1 + (num_commas_e15 // 3))
        # Use the max of B15 and E15 line counts for row 15
        before_height_15 = ws.row_dimensions[15].height
        final_height_15 = max(before_height_15, num_lines_e15 * 20)
        ws.row_dimensions[15].height = final_height_15
        logger.info(f"[EXCEL] Row 15 E15 (asset): {num_lines_e15} lines ({num_commas_e15} commas), before={before_height_15}, after={final_height_15}")

        # End dates with dynamic row height
        end_dates_value = get_end_dates()
        ws["E17"] = end_dates_value
        ws["E17"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        num_lines_e17 = end_dates_value.count('\n') + 1 if end_dates_value else 1
        # Use the max of B17 and E17 line counts for row 17
        before_height_17 = ws.row_dimensions[17].height
        final_height_17 = max(before_height_17, num_lines_e17 * 15)
        ws.row_dimensions[17].height = final_height_17
        logger.info(f"[EXCEL] Row 17 E17 (end dates): {num_lines_e17} lines, before={before_height_17}, after={final_height_17}")

        # Category in E19 - also set wrap text and check if it needs more height
        category_value = format_value(data.get("category"))
        ws["E19"] = category_value
        ws["E19"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        num_lines_e19 = category_value.count('\n') + 1 if category_value else 1
        # Use the max of B19 and E19 line counts for row 19
        before_height_19 = ws.row_dimensions[19].height
        final_height_19 = max(before_height_19, num_lines_e19 * 20)
        ws.row_dimensions[19].height = final_height_19
        logger.info(f"[EXCEL] Row 19 E19 (category): {num_lines_e19} lines, before={before_height_19}, after={final_height_19}")

        # SLA - show the deduction amount, not the percentage
        # E.g., if 10% of 462,520 = 46,252, show 46,252
        sla_deduction = data.get("sla_deduction", 0) or 0
        ws["E21"] = sla_deduction
        ws["E21"].number_format = excel_number_format
        ws["E23"] = data.get("municipality_fee", 0)                     # DM (Dubai Municipality)
        ws["E23"].number_format = excel_number_format

        # Payment terms with dynamic row height
        payment_terms_value = format_value(data.get("payment_terms"))
        ws["E25"] = payment_terms_value
        ws["E25"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        num_lines_e25 = payment_terms_value.count('\n') + 1 if payment_terms_value else 1
        ws.row_dimensions[25].height = max(15, num_lines_e25 * 15)

        ws["E27"] = format_value(data.get("sales_person"))              # Sales Person Name
        ws["E29"] = data.get("commission_pct", 0)                       # Commission%

        # Head of Sales Signature section
        ws["A39"] = "Head of Sales Signature"

        # Add HoS signature if provided (in italics)
        hos_signature = data.get("hos_signature")
        if hos_signature:
            ws["B39"] = hos_signature
            ws["B39"].font = openpyxl.styles.Font(italic=True)

        # Net rentals excl SLA: write value to B29 and breakdown to A33
        sla_pct = data.get("sla_pct", 0) or 0

        # Write net excl SLA value directly to cell B29 with currency formatting
        ws['B29'] = net_rentals_excl_sla
        ws['B29'].number_format = excel_number_format

        # Always show location breakdown in A33 with numbered list
        locations = data.get("locations", [])
        location_lines = ["Location splits excl SLA:"]

        for idx, location in enumerate(locations, 1):
            loc_name = location.get("name", "Unknown")
            # If SLA was applied, use post_sla_amount; otherwise use net_amount
            if sla_pct > 0:
                amount = location.get("post_sla_amount", 0)
            else:
                amount = location.get("net_amount", 0)

            if amount > 0:
                location_lines.append(f"{idx}. {loc_name} - {amount:,.2f}")

        # Write location splits to A33 (merged cell)
        if len(location_lines) > 1:
            location_splits_text = "\n".join(location_lines)

            for merged_range in ws.merged_cells.ranges:
                if "A33" in merged_range:
                    # Get the top-left cell of the merged range
                    min_col, min_row = merged_range.bounds[0], merged_range.bounds[1]
                    top_left_cell = ws.cell(row=min_row, column=min_col)
                    top_left_cell.value = location_splits_text
                    top_left_cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                    # Remove underline formatting
                    top_left_cell.font = openpyxl.styles.Font(
                        name=top_left_cell.font.name,
                        size=top_left_cell.font.size,
                        bold=top_left_cell.font.bold,
                        italic=top_left_cell.font.italic,
                        underline='none',
                        color=top_left_cell.font.color
                    )
                    break

        # Save to temporary file
        temp_excel = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_excel_path = Path(temp_excel.name)
        wb.save(temp_excel_path)
        logger.info(f"[BOOKING PARSER] Generated branded Excel: {temp_excel_path}")

        return temp_excel_path

    async def generate_combined_pdf(self, data: dict[str, Any], bo_ref: str, original_bo_path: Path, apply_stamp: bool = False) -> Path:
        """
        Generate combined PDF: Excel (converted to PDF) + Original BO PDF concatenated

        Args:
            data: Parsed booking order data
            bo_ref: BO reference for filename
            original_bo_path: Path to the original BO file (PDF or will be converted)
            apply_stamp: If True, applies HoS approval stamp to original BO (only on HoS approval)

        Returns:
            Path to the combined PDF file
        """
        logger.info(f"[BOOKING PARSER] Generating combined PDF for {bo_ref}, original_bo_path: {original_bo_path} (type: {type(original_bo_path)}), apply_stamp: {apply_stamp}")

        excel_path = None
        excel_pdf_path = None
        original_pdf_path = None
        stamped_original_pdf_path = None

        try:
            # Step 1: Generate Excel
            logger.info(f"[BOOKING PARSER] Step 1: Generating Excel for {bo_ref}")
            excel_path = await self.generate_excel(data, bo_ref)
            logger.info(f"[BOOKING PARSER] Excel generated: {excel_path}")

            # Step 2: Convert Excel to PDF using LibreOffice
            logger.info("[BOOKING PARSER] Step 2: Converting Excel to PDF")
            excel_pdf_path = await self._convert_excel_to_pdf(excel_path)
            logger.info(f"[BOOKING PARSER] Excel PDF created: {excel_pdf_path}")

            # Step 3: Ensure original BO is PDF (convert if needed)
            logger.info(f"[BOOKING PARSER] Step 3: Ensuring original BO is PDF, path: {original_bo_path}")
            original_pdf_path = await self._ensure_pdf(original_bo_path)
            logger.info(f"[BOOKING PARSER] Original BO PDF ready: {original_pdf_path}")

            # Step 3.5: Apply stamp to original BO PDF (only if HoS approved)
            if apply_stamp:
                logger.info("[BOOKING PARSER] Step 3.5: Applying HoS approval stamp to original BO")
                stamped_original_pdf_path = await self._apply_stamp_to_pdf(original_pdf_path)
                logger.info(f"[BOOKING PARSER] Stamped original BO ready: {stamped_original_pdf_path}")
            else:
                logger.info("[BOOKING PARSER] Skipping stamp application (not HoS approved)")
                stamped_original_pdf_path = original_pdf_path

            # Step 4: Concatenate PDFs (Excel PDF first, then original BO)
            combined_pdf_path = COMBINED_BOS_DIR / f"{bo_ref}_combined.pdf"
            await self._concatenate_pdfs([excel_pdf_path, stamped_original_pdf_path], combined_pdf_path)

            logger.info(f"[BOOKING PARSER] Combined PDF generated: {combined_pdf_path}")
            return combined_pdf_path

        finally:
            # Clean up temporary files (always runs, even if there's an error)
            try:
                if excel_path and excel_path.exists():
                    excel_path.unlink()
                if excel_pdf_path and excel_pdf_path != excel_path and excel_pdf_path.exists():
                    excel_pdf_path.unlink()
                if original_pdf_path and original_pdf_path != original_bo_path and original_pdf_path.exists():
                    original_pdf_path.unlink()
                if stamped_original_pdf_path and stamped_original_pdf_path != original_pdf_path and stamped_original_pdf_path.exists():
                    stamped_original_pdf_path.unlink()
            except Exception as e:
                logger.warning(f"[BOOKING PARSER] Failed to clean up temp files: {e}")

    async def _convert_excel_to_pdf(self, excel_path: Path) -> Path:
        """Convert Excel file to PDF using LibreOffice"""
        logger.info(f"[BOOKING PARSER] Converting Excel to PDF: {excel_path}")

        output_dir = excel_path.parent
        pdf_path = excel_path.with_suffix('.pdf')

        try:
            # Use LibreOffice to convert Excel to PDF
            result = subprocess.run([
                'soffice',
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', str(output_dir),
                str(excel_path)
            ], capture_output=True, text=True, timeout=30)

            if result.returncode != 0:
                raise Exception(f"LibreOffice conversion failed: {result.stderr}")

            if not pdf_path.exists():
                raise Exception(f"PDF not created at expected path: {pdf_path}")

            logger.info(f"[BOOKING PARSER] Excel converted to PDF: {pdf_path}")
            return pdf_path

        except Exception as e:
            logger.error(f"[BOOKING PARSER] Failed to convert Excel to PDF: {e}")
            raise

    async def _ensure_pdf(self, file_path: Path) -> Path:
        """Ensure file is PDF, convert if needed"""
        if file_path.suffix.lower() == '.pdf':
            return file_path

        logger.info(f"[BOOKING PARSER] Converting {file_path.suffix} to PDF")

        output_dir = file_path.parent
        pdf_path = file_path.with_suffix('.pdf')

        try:
            result = subprocess.run([
                'soffice',
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', str(output_dir),
                str(file_path)
            ], capture_output=True, text=True, timeout=30)

            if result.returncode != 0:
                raise Exception(f"LibreOffice conversion failed: {result.stderr}")

            if not pdf_path.exists():
                raise Exception(f"PDF not created at expected path: {pdf_path}")

            logger.info(f"[BOOKING PARSER] File converted to PDF: {pdf_path}")
            return pdf_path

        except Exception as e:
            logger.error(f"[BOOKING PARSER] Failed to convert to PDF: {e}")
            raise

    async def _apply_stamp_to_pdf(self, pdf_path: Path) -> Path:
        """
        Apply stamp.png to the PDF document using intelligent placement algorithm
        that avoids overlapping with existing content.

        Algorithm:
        1. Renders PDF page to image and builds ink mask (text/graphics detection)
        2. Searches corners in priority order (BR, BL, TR, TL) for empty space
        3. If no space found, tries smaller stamp sizes (down to 85% of original)
        4. Fallback: uses distance transform to find largest empty area

        Returns path to stamped PDF
        """
        logger.info(f"[STAMP] Applying stamp to PDF with smart placement: {pdf_path}")

        stamp_img_path = config.BASE_DIR / "stamp.png"
        if not stamp_img_path.exists():
            logger.warning(f"[STAMP] stamp.png not found at {stamp_img_path}, skipping stamp")
            return pdf_path

        try:
            import io

            import cv2
            import fitz  # PyMuPDF
            import numpy as np
            from PIL import Image
            from pypdf import PdfReader, PdfWriter
            from reportlab.lib.utils import ImageReader
            from reportlab.pdfgen import canvas

            # Helper functions for smart placement
            def mm_to_in(mm):
                return mm / 25.4

            def px_to_pt(px, dpi):
                return px / dpi * 72.0

            def integral_sum(ii, x, y, w, h):
                """Fast sum of rectangle using integral image"""
                x2, y2 = x + w, y + h
                return ii[y2, x2] - ii[y, x2] - ii[y2, x] + ii[y, x]

            def build_ink_mask(gray, stamp_size_mm):
                """Build robust ink mask from grayscale image with clearance scaled to stamp size"""
                try:
                    # Background flattening to handle uneven lighting
                    bg = cv2.GaussianBlur(gray, (0, 0), sigmaX=21, sigmaY=21)
                    # Prevent divide by zero
                    bg = np.maximum(bg, 1)
                    norm = cv2.divide(gray, bg, scale=128)

                    # Adaptive threshold for local text detection
                    bw = cv2.adaptiveThreshold(
                        norm, 255,
                        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                        cv2.THRESH_BINARY, 41, 8
                    )
                    ink = (bw == 0).astype(np.uint8)

                    # Remove noise/speckles
                    n, labels, stats, _ = cv2.connectedComponentsWithStats(ink, connectivity=8)
                    min_area = 30
                    keep = np.zeros_like(ink)
                    for i in range(1, n):
                        if stats[i, cv2.CC_STAT_AREA] >= min_area:
                            keep[labels == i] = 1
                    ink = keep

                    # Add safety buffer around text - scale clearance based on stamp size
                    # Base: 55mm stamp -> 9x9 kernel, 10 iterations
                    # Scales DOWN as stamp gets smaller (smaller stamps need less clearance)
                    scale_factor = stamp_size_mm / 55.0
                    kernel_size = max(3, int(9 * scale_factor))
                    # Ensure odd number for kernel
                    if kernel_size % 2 == 0:
                        kernel_size += 1
                    iterations = max(4, int(10 * scale_factor))

                    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (kernel_size, kernel_size))
                    ink = cv2.dilate(ink, kernel, iterations=iterations)
                    logger.debug(f"[STAMP] Clearance buffer: {kernel_size}x{kernel_size} kernel, {iterations} iterations for {stamp_size_mm}mm stamp")
                    return ink
                except Exception as e:
                    logger.warning(f"[STAMP] Error building ink mask: {e}, using empty mask")
                    # Return empty mask (all white) if processing fails
                    return np.zeros_like(gray, dtype=np.uint8)

            def scan_corner(integral, W, H, ww, wh, corner, margin, stride, max_ink_ratio):
                """Scan a corner for empty space"""
                x_min, y_min = margin, margin
                x_max, y_max = W - margin - ww, H - margin - wh
                if x_max < x_min or y_max < y_min:
                    return None

                # Integral image is (H+1, W+1), so max valid indices are H, W
                integral_h, integral_w = integral.shape

                # Set scan direction based on corner
                if corner == "BR":
                    ys = range(y_max, y_min - 1, -stride)
                    xs = range(x_max, x_min - 1, -stride)
                elif corner == "BL":
                    ys = range(y_max, y_min - 1, -stride)
                    xs = range(x_min, x_max + 1, stride)
                elif corner == "TR":
                    ys = range(y_min, y_max + 1, stride)
                    xs = range(x_max, x_min - 1, -stride)
                else:  # "TL"
                    ys = range(y_min, y_max + 1, stride)
                    xs = range(x_min, x_max + 1, stride)

                area = ww * wh
                for y in ys:
                    for x in xs:
                        # Bounds check: ensure x+ww and y+wh don't exceed integral bounds
                        if x + ww < integral_w and y + wh < integral_h:
                            s = integral_sum(integral, x, y, ww, wh)
                            if s / area <= max_ink_ratio:
                                return (x, y)
                return None

            def find_spot(ink_masks, stamp_w_px, stamp_h_px, margin_px, corner_order,
                          stride_px, max_ink_ratio, min_scale, scale_step):
                """Find optimal placement for stamp with scaled clearance"""
                H, W = ink_masks[1.0].shape
                found = None
                used_w = used_h = None

                # Try multiple stamp sizes (with max iteration guard)
                scale = 1.0
                max_iterations = 25  # Increased to try more size combinations
                iteration = 0
                while found is None and scale >= min_scale and iteration < max_iterations:
                    ww = max(8, int(round(stamp_w_px * scale)))
                    wh = max(8, int(round(stamp_h_px * scale)))

                    # Select appropriate ink mask based on scale
                    # Use the closest pre-computed mask
                    if scale >= 0.85:
                        ink = ink_masks[1.0]  # Large stamp, large clearance
                    elif scale >= 0.70:
                        ink = ink_masks[0.85]  # Medium-large stamp
                    elif scale >= 0.60:
                        ink = ink_masks[0.70]  # Medium stamp
                    else:
                        ink = ink_masks[0.50]  # Small stamp, small clearance

                    integral = cv2.integral(ink)

                    for corner in corner_order:
                        pt = scan_corner(integral, W, H, ww, wh, corner, margin_px, stride_px, max_ink_ratio)
                        if pt:
                            found = (pt[0], pt[1])
                            used_w, used_h = ww, wh
                            logger.info(f"[STAMP] Found spot in {corner} corner at scale {scale:.2f}")
                            break

                    if not found:
                        scale *= scale_step
                        iteration += 1

                # Fallback: distance transform to find largest empty area
                if found is None:
                    logger.info("[STAMP] No corner worked, using distance transform fallback")
                    try:
                        # Use the smallest ink mask (least clearance) for fallback
                        ink_fallback = ink_masks[0.50]
                        integral_fallback = cv2.integral(ink_fallback)
                        bg = (ink_fallback == 0).astype(np.uint8)
                        dist = cv2.distanceTransform(bg, cv2.DIST_L2, 5)

                        # Try progressively smaller sizes and higher ink tolerance
                        for fallback_scale in [1.0, 0.8, 0.6, 0.5]:
                            ww0 = max(8, int(round(stamp_w_px * fallback_scale)))
                            wh0 = max(8, int(round(stamp_h_px * fallback_scale)))
                            half_w, half_h = max(1, ww0 // 2), max(1, wh0 // 2)

                            # Ensure valid region is within bounds
                            if half_h < H and half_w < W:
                                valid = np.zeros_like(dist, dtype=bool)
                                valid[half_h:H - half_h, half_w:W - half_w] = True
                                dist_masked = np.where(valid, dist, 0)
                                y0, x0 = np.unravel_index(np.argmax(dist_masked), dist_masked.shape)
                                x = int(x0 - ww0 // 2)
                                y = int(y0 - wh0 // 2)
                                if x >= 0 and y >= 0 and x + ww0 <= W and y + wh0 <= H:
                                    s = integral_sum(integral_fallback, x, y, ww0, wh0)
                                    # Accept up to 20% ink for fallback (more lenient)
                                    if s / (ww0 * wh0) <= 0.20:
                                        found = (x, y)
                                        used_w, used_h = ww0, wh0
                                        logger.info(f"[STAMP] Distance transform found spot at scale {fallback_scale:.2f}")
                                        break

                        # Last resort: just place it at bottom-right corner regardless
                        if found is None:
                            logger.warning("[STAMP] All fallbacks failed, forcing placement at bottom-right")
                            ww0 = max(8, int(round(stamp_w_px * 0.4)))  # 40% size
                            wh0 = max(8, int(round(stamp_h_px * 0.4)))
                            x = max(0, W - ww0 - margin_px)
                            y = max(0, H - wh0 - margin_px)
                            if x >= 0 and y >= 0:
                                found = (x, y)
                                used_w, used_h = ww0, wh0
                                logger.info(f"[STAMP] Forced placement at ({x}, {y}) with size {ww0}x{wh0}")
                    except Exception as e:
                        logger.warning(f"[STAMP] Distance transform fallback failed: {e}")

                return (found, used_w, used_h)

            # Configuration
            stamp_width_mm = 55.0  # 55mm = ~2.17 inches (start smaller for better fit)
            dpi = 200
            margin_mm = 15.0  # 15mm (~0.6 inches) margin from page edges
            stride_px = 12
            max_ink_ratio = 0.10  # Tolerate 10% ink in region (allows some overlap)
            corner_order = ("BR", "BL", "TR", "TL")
            min_scale = 0.40  # Try down to 40% of original size (22mm minimum)
            scale_step = 0.98  # 2% reduction per attempt (more granular steps)

            # Load stamp image and add date
            stamp_img = Image.open(stamp_img_path).convert("RGBA")

            # Add today's date to the original stamp
            dated_stamp_path = None
            try:
                import tempfile
                from datetime import datetime

                from PIL import ImageDraw, ImageFont

                # Get today's date in DD-MM-YYYY format
                today = datetime.now().strftime("%d-%m-%Y")

                # Create drawing context on the logo itself (don't extend canvas)
                draw = ImageDraw.Draw(stamp_img)

                # Get stamp dimensions
                width, height = stamp_img.size

                # Try to use a nice BOLD font, fallback to default
                # Use EXACT same code as test_stamp.py for consistency
                try:
                    # Use Helvetica Bold
                    font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 60, index=1)  # index=1 for bold
                except OSError:
                    try:
                        font = ImageFont.truetype("/Library/Fonts/Arial Bold.ttf", 48)
                    except OSError:
                        try:
                            font = ImageFont.truetype("/Library/Fonts/Arial.ttf", 48)
                        except OSError:
                            logger.debug("[BO PARSER] Using default font for stamp - system fonts not available")
                            font = ImageFont.load_default()

                # Position date lower on the stamp
                # Get text bounding box to center it
                bbox = draw.textbbox((0, 0), today, font=font)
                text_width = bbox[2] - bbox[0]
                text_x = (width - text_width) // 2 + 40  # Shifted right by 40 pixels
                text_y = int(height * 0.521)  # Place at 52.1% down from top

                # Draw date in black with white outline for visibility
                # Draw outline (white)
                for adj_x in [-2, -1, 0, 1, 2]:
                    for adj_y in [-2, -1, 0, 1, 2]:
                        if adj_x != 0 or adj_y != 0:
                            draw.text((text_x + adj_x, text_y + adj_y), today, fill=(255, 255, 255, 255), font=font)
                # Draw main text (black)
                draw.text((text_x, text_y), today, fill=(0, 0, 0, 255), font=font)

                # Save dated stamp to temporary file
                temp_fd, dated_stamp_path = tempfile.mkstemp(suffix='.png')
                os.close(temp_fd)
                stamp_img.save(dated_stamp_path)

                logger.info(f"[STAMP] Added date {today} to stamp, saved to temp file")
            except Exception as e:
                logger.warning(f"[STAMP] Failed to add date to stamp: {e}, using stamp without date")
                dated_stamp_path = None

            # Use dated stamp if available, otherwise original
            final_stamp_path = dated_stamp_path if dated_stamp_path else stamp_img_path

            stamp_width, stamp_height = stamp_img.size
            stamp_aspect_ratio = stamp_width / stamp_height

            # Calculate stamp size in pixels
            w_in = mm_to_in(stamp_width_mm)
            h_in = w_in / stamp_aspect_ratio
            stamp_w_px = int(round(w_in * dpi))
            stamp_h_px = int(round(h_in * dpi))
            margin_px = int(round(mm_to_in(margin_mm) * dpi))

            # Open PDF with PyMuPDF and render first page
            doc = fitz.open(str(pdf_path))
            try:
                page = doc[0]

                # Render page to image for analysis
                zoom = dpi / 72.0
                mat = fitz.Matrix(zoom, zoom)
                pix = page.get_pixmap(matrix=mat, alpha=False)

                # Validate pixmap dimensions
                if pix.h <= 0 or pix.w <= 0:
                    logger.warning(f"[STAMP] Invalid page dimensions: {pix.w}x{pix.h}, skipping stamp")
                    doc.close()
                    return pdf_path

                img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
                if pix.n == 1:
                    img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
                gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

                # Build multiple ink masks with different clearance levels
                # Larger stamps need more clearance, smaller stamps need less
                logger.info("[STAMP] Building ink masks with scaled clearance for content detection")
                ink_masks = {}
                for scale_level in [1.0, 0.85, 0.70, 0.50]:
                    scaled_size = stamp_width_mm * scale_level
                    ink_masks[scale_level] = build_ink_mask(gray, scaled_size)
                logger.info(f"[STAMP] Created {len(ink_masks)} ink masks for different stamp sizes")
            except Exception as e:
                logger.warning(f"[STAMP] Failed to render page: {e}, skipping stamp")
                doc.close()
                return pdf_path

            # Find optimal placement
            logger.info("[STAMP] Searching for optimal stamp placement")
            (found, ww, wh) = find_spot(
                ink_masks, stamp_w_px, stamp_h_px, margin_px, corner_order,
                stride_px, max_ink_ratio, min_scale, scale_step
            )

            if not found:
                logger.warning("[STAMP] Could not find suitable placement, skipping stamp")
                doc.close()
                return pdf_path

            x_px, y_px = found

            # Convert pixel coordinates to PDF points
            page_width = float(page.rect.width)
            page_height = float(page.rect.height)
            w_pt = px_to_pt(ww, dpi)
            h_pt = px_to_pt(wh, dpi)
            x_pt = x_px / zoom
            y_pt_from_top = y_px / zoom
            y_pt = page_height - (y_pt_from_top + h_pt)

            logger.info(f"[STAMP] Placing stamp at ({x_pt:.1f}, {y_pt:.1f}), size: {w_pt:.1f}x{h_pt:.1f} pt")

            doc.close()

            # Now use pypdf to add the stamp
            reader = PdfReader(str(pdf_path))
            writer = PdfWriter()

            # Create overlay PDF with stamp
            packet = io.BytesIO()
            first_page = reader.pages[0]
            can = canvas.Canvas(packet, pagesize=(float(first_page.mediabox.width),
                                                   float(first_page.mediabox.height)))

            # Draw stamp at calculated position (use dated stamp file)
            can.drawImage(ImageReader(final_stamp_path), x_pt, y_pt,
                         width=w_pt, height=h_pt,
                         mask='auto', preserveAspectRatio=False)
            can.save()

            # Merge stamp onto first page
            packet.seek(0)
            stamp_pdf = PdfReader(packet)

            for idx, page in enumerate(reader.pages):
                if idx == 0:
                    page.merge_page(stamp_pdf.pages[0])
                writer.add_page(page)

            # Write output
            output_path = pdf_path.parent / f"stamped_{pdf_path.name}"
            try:
                with open(output_path, 'wb') as output_file:
                    writer.write(output_file)
                logger.info(f"[STAMP] Successfully applied stamp to {output_path}")

                # Clean up temp dated stamp file
                if dated_stamp_path and os.path.exists(dated_stamp_path):
                    try:
                        os.unlink(dated_stamp_path)
                    except OSError as cleanup_err:
                        logger.warning(f"[STAMP] Failed to cleanup temp stamp file {dated_stamp_path}: {cleanup_err}")

                return output_path
            except Exception as write_error:
                logger.error(f"[STAMP] Failed to write stamped PDF: {write_error}")
                # Clean up partial file if write failed
                if output_path.exists():
                    try:
                        output_path.unlink()
                    except OSError as cleanup_err:
                        logger.warning(f"[STAMP] Failed to cleanup partial output file: {cleanup_err}")
                # Clean up temp dated stamp file
                if dated_stamp_path and os.path.exists(dated_stamp_path):
                    try:
                        os.unlink(dated_stamp_path)
                    except OSError as cleanup_err:
                        logger.warning(f"[STAMP] Failed to cleanup temp stamp file: {cleanup_err}")
                raise write_error

        except Exception as e:
            logger.error(f"[STAMP] Failed to apply stamp: {e}")
            logger.exception(e)
            # Return original PDF if stamping fails
            return pdf_path

    async def _concatenate_pdfs(self, pdf_paths: list[Path], output_path: Path) -> None:
        """Concatenate multiple PDFs into one"""
        logger.info(f"[BOOKING PARSER] Concatenating {len(pdf_paths)} PDFs")

        try:
            merger = PdfMerger()

            for pdf_path in pdf_paths:
                logger.info(f"[BOOKING PARSER] Adding PDF: {pdf_path}")
                merger.append(str(pdf_path))

            merger.write(str(output_path))
            merger.close()

            logger.info(f"[BOOKING PARSER] PDFs concatenated successfully: {output_path}")

        except Exception as e:
            logger.error(f"[BOOKING PARSER] Failed to concatenate PDFs: {e}")
            raise
